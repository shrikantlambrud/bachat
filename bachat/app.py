import os
from flask import Flask, render_template, request, redirect, url_for, session, flash, jsonify, send_file
from datetime import datetime, date, timedelta
import mysql.connector
from werkzeug.security import generate_password_hash, check_password_hash
import functools
from decimal import Decimal, ROUND_HALF_UP  # Import Decimal for precise arithmetic and rounding
import json  # Import json for storing disbursement details
from io import BytesIO  # For potential in-memory file handling, though direct file generation is limited

# Attempt to import report generation libraries. These might not be available in all environments.
# If they are not installed, the application will still run, but export functionality will be conceptual.
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    EXCEL_AVAILABLE = True
except ImportError:
    print("openpyxl not found. Excel export will be conceptual.")
    EXCEL_AVAILABLE = False

try:
    from reportlab.lib.pagesizes import letter
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib import colors

    PDF_AVAILABLE = True
except ImportError:
    print("reportlab not found. PDF export will be conceptual.")
    PDF_AVAILABLE = False

# Import configuration from config.py
from config import Config

app = Flask(__name__)
app.config.from_object(Config)  # Load configuration from Config class

# Database configuration is now loaded from app.config
DB_CONFIG = app.config['DB_CONFIG']


# Context Processor to make datetime available in all templates
@app.context_processor
def inject_datetime():
    """Makes the datetime object and date object available globally in all Jinja2 templates."""
    return dict(datetime=datetime, date=date)

@app.template_filter('from_json')
def from_json_filter(value):
    """Custom Jinja2 filter to parse JSON strings."""
    if value:
        try:
            return json.loads(value)
        except json.JSONDecodeError:
            return {}
    return {}


def get_db_connection():
    """Establishes a connection to the MySQL database."""
    try:
        conn = mysql.connector.connect(**DB_CONFIG)
        return conn
    except mysql.connector.Error as err:
        print(f"Error connecting to database: {err}")
        return None


# Modified login_required decorator to accept a list of roles
def login_required(roles=None):
    """Decorator to ensure user is logged in and has the required role(s)."""

    def wrapper(view_func):
        @functools.wraps(view_func)
        def decorated_function(*args, **kwargs):
            if 'user_id' not in session:
                flash('Please log in to access this page.', 'danger')
                return redirect(url_for('login'))

            conn = get_db_connection()
            if conn is None:
                flash('Database connection error. Please try again later.', 'danger')
                return redirect(url_for('login'))

            cursor = conn.cursor(dictionary=True, buffered=True)
            cursor.execute("SELECT role FROM users WHERE id = %s", (session['user_id'],))
            user = cursor.fetchone()
            cursor.close()  # Close cursor immediately after fetching
            conn.close()

            if not user:
                session.pop('user_id', None)
                flash('User not found. Please log in again.', 'danger')
                return redirect(url_for('login'))

            # Check if roles are specified and if the user's role is in the allowed roles
            if roles and user['role'] not in roles:
                # Construct a user-friendly message for allowed roles
                allowed_roles_str = ", ".join([r.capitalize() for r in roles])
                flash(
                    f'Access denied. You need to be one of the following roles: {allowed_roles_str} to view this page.',
                    'danger')
                return redirect(url_for('dashboard'))  # Redirect to their own dashboard

            return view_func(*args, **kwargs)

        return decorated_function

    return wrapper


@app.route('/')
def index():
    """Redirects to the login page."""
    return redirect(url_for('login'))


@app.route('/login', methods=['GET', 'POST'])
def login():
    """Handles user login."""
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']

        conn = get_db_connection()
        if conn is None:
            flash('Database connection error.', 'danger')
            return render_template('login.html')

        # Use buffered=True for the cursor
        cursor = conn.cursor(dictionary=True, buffered=True)
        cursor.execute("SELECT id, username, password, role, name FROM users WHERE username = %s", (username,))
        user = cursor.fetchone()
        cursor.close()  # Close cursor immediately after fetching
        conn.close()

        # Added check to ensure user['password'] is not None or an empty string before hashing
        if user and user['password'] and user['password'].strip() and check_password_hash(user['password'], password):
            session['user_id'] = user['id']
            session['username'] = user['username']
            session['role'] = user['role']
            session['name'] = user['name']
            flash(f'Welcome, {user["name"]}!', 'success')
            return redirect(url_for('dashboard'))
        else:
            flash('Invalid username or password.', 'danger')
    return render_template('login.html')


@app.route('/logout')
def logout():
    """Logs out the current user."""
    session.pop('user_id', None)
    session.pop('username', None)
    session.pop('role', None)
    session.pop('name', None)
    flash('You have been logged out.', 'info')
    return redirect(url_for('login'))


@app.route('/register', methods=['GET', 'POST'])
def register():
    """Handles user registration (initially for President, then President can add members)."""
    if request.method == 'POST':
        name = request.form['name']
        username = request.form['username']
        email = request.form['email']
        contact_number = request.form['contact_number']
        pan_number = request.form['pan_number']
        aadhar_number = request.form['aadhar_number']
        password = request.form['password']
        role = request.form.get('role', 'member')  # Default to 'member' if not specified

        # Added server-side validation for password
        if not password or not password.strip():  # Ensure password is not empty or just whitespace
            flash('Password cannot be empty.', 'danger')
            return render_template('register.html')

        hashed_password = generate_password_hash(password)

        conn = get_db_connection()
        if conn is None:
            flash('Database connection error.', 'danger')
            return render_template('register.html')

        # Use buffered=True for the cursor
        cursor = conn.cursor(buffered=True)
        try:
            cursor.execute(
                "INSERT INTO users (name, username, email, contact_number, pan_number, aadhar_number, password, role) VALUES (%s, %s, %s, %s, %s, %s, %s, %s)",
                (name, username, email, contact_number, pan_number, aadhar_number, hashed_password, role)
            )
            conn.commit()
            flash('Registration successful! Please log in.', 'success')
            return redirect(url_for('login'))
        except mysql.connector.Error as err:
            if err.errno == 1062:  # Duplicate entry error
                flash('Username, Email, PAN, or Aadhar number already exists.', 'danger')
            else:
                flash(f'An error occurred during registration: {err}', 'danger')
            conn.rollback()
        finally:
            cursor.close()
            conn.close()
    return render_template('register.html')


@app.route('/dashboard')
@login_required()
def dashboard():
    """Renders the appropriate dashboard based on user role."""
    role = session.get('role')
    user_id = session.get('user_id')
    conn = get_db_connection()
    if conn is None:
        flash('Database connection error. Please try again later.', 'danger')
        return redirect(url_for('logout'))  # Redirect to logout if DB connection fails
    # Use buffered=True for the cursor
    cursor = conn.cursor(dictionary=True, buffered=True)

    # Initialize variables for president/secretary dashboard with default values
    total_members = 0
    total_loans = 0
    total_contributions_this_month = Decimal('0.00')
    total_interest_this_month = Decimal('0.00')
    total_income_this_month = Decimal('0.00')
    bank_balance = Decimal('0.00')
    recent_activities = []
    pending_contributions_count = 0  # New: for President/Secretary dashboard

    # Initialize variables for member dashboard with default values (if needed for a different path)
    total_contributed = Decimal('0.00')
    current_loans = []
    pending_contribution = None

    # Modified to allow both president and secretary to view the president dashboard
    if role in ['president', 'secretary']:
        # Get total members (count all users)
        cursor.execute("SELECT COUNT(*) as total_members FROM users")
        total_members = cursor.fetchone()['total_members']

        # Get total loans (approved)
        cursor.execute("SELECT COUNT(*) as total_loans FROM loans WHERE status = 'approved'")
        total_loans = cursor.fetchone()['total_loans']

        # Get total contributions this month
        current_month = datetime.now().month
        current_year = datetime.now().year
        cursor.execute(
            "SELECT SUM(amount) as total_contributions FROM contributions WHERE month = %s AND year = %s AND is_paid = TRUE",
            (current_month, current_year))
        result_contributions = cursor.fetchone()
        total_contributions_this_month = result_contributions['total_contributions'] if result_contributions and \
                                                                                        result_contributions[
                                                                                            'total_contributions'] is not None else Decimal(
            '0.00')

        # Get total interest collected this month
        cursor.execute("""
            SELECT SUM(lp.interest_paid) as total_interest_paid 
            FROM loan_payments lp
            JOIN loans l ON lp.loan_id = l.id
            WHERE MONTH(lp.payment_date) = %s AND YEAR(lp.payment_date) = %s
        """, (current_month, current_year))
        result_interest = cursor.fetchone()
        total_interest_this_month = result_interest['total_interest_paid'] if result_interest and result_interest[
            'total_interest_paid'] is not None else Decimal('0.00')

        # Calculate total income for this month
        total_income_this_month = total_contributions_this_month + total_interest_this_month

        # Get bank balance
        cursor.execute("SELECT balance FROM bank_balance WHERE id = 1")
        result_balance = cursor.fetchone()
        bank_balance = result_balance['balance'] if result_balance and result_balance[
            'balance'] is not None else Decimal('0.00')

        # Get recent activities (e.g., last 5 contributions/loans)
        cursor.execute("""
            SELECT 'contribution' as type, u.name as member_name, c.amount, c.payment_date 
            FROM contributions c JOIN users u ON c.user_id = u.id 
            WHERE c.is_paid = TRUE ORDER BY c.payment_date DESC LIMIT 5
        """)
        recent_contributions = cursor.fetchall()

        cursor.execute("""
            SELECT 'loan' as type, u.name as member_name, l.amount, l.start_date 
            FROM loans l JOIN users u ON l.user_id = u.id 
            ORDER BY l.start_date DESC LIMIT 5
        """)
        recent_loans = cursor.fetchall()

        # Convert all dates to date objects for consistent comparison before sorting
        for activity in recent_contributions:
            if 'payment_date' in activity and isinstance(activity['payment_date'], datetime):
                activity['payment_date'] = activity['payment_date'].date()
        for activity in recent_loans:
            if 'start_date' in activity and isinstance(activity['start_date'], datetime):
                activity['start_date'] = activity['start_date'].date()

        recent_activities = sorted(recent_contributions + recent_loans,
                                   key=lambda x: x['payment_date'] if 'payment_date' in x else x['start_date'],
                                   reverse=True)[:5]

        cursor.close()
        conn.close()
        return render_template('president_dashboard.html',
                               total_members=total_members,
                               total_loans=total_loans,
                               total_contributions_this_month=total_contributions_this_month,  # Updated variable name
                               total_interest_this_month=total_interest_this_month,  # New variable
                               total_income_this_month=total_income_this_month,  # New variable
                               bank_balance=bank_balance,
                               recent_activities=recent_activities,
                               pending_contributions_count=pending_contributions_count)  # Pass pending count

    elif role == 'member':  # Only 'member' role for this block now
        # Member dashboard
        # Get user's total contributions
        cursor.execute(
            "SELECT SUM(amount) as total_contributed FROM contributions WHERE user_id = %s AND is_paid = TRUE",
            (user_id,))
        result_contributed = cursor.fetchone()
        total_contributed = result_contributed['total_contributed'] if result_contributed and result_contributed[
            'total_contributed'] is not None else Decimal('0.00')

        # Get user's current loans
        cursor.execute("SELECT * FROM loans WHERE user_id = %s AND status IN ('approved', 'overdue')", (user_id,))
        current_loans = cursor.fetchall()

        # Get user's pending contributions for current month
        current_month = datetime.now().month
        current_year = datetime.now().year
        cursor.execute("""
            SELECT * FROM contributions 
            WHERE user_id = %s AND month = %s AND year = %s AND is_paid = FALSE
        """, (user_id, current_month, current_year))
        pending_contribution = cursor.fetchone()

        cursor.close()
        conn.close()
        return render_template('member_dashboard.html',
                               total_contributed=total_contributed,
                               current_loans=current_loans,
                               pending_contribution=pending_contribution,
                               is_president=(role == 'president'))  # Pass this for conditional rendering

    else:
        flash('Unknown role.', 'danger')
        cursor.close()
        conn.close()
        return redirect(url_for('login'))


@app.route('/members')
@login_required(roles=['president', 'secretary'])  # Allow secretary to manage members
def manage_members():
    """Displays a list of all members for the President/Secretary to manage."""
    conn = get_db_connection()
    if conn is None:
        flash('Database connection error. Please try again later.', 'danger')
        return redirect(url_for('manage_members'))
    # Use buffered=True for the cursor
    cursor = conn.cursor(dictionary=True, buffered=True)
    cursor.execute("SELECT id, name, username, email, contact_number, role FROM users ORDER BY id")
    members = cursor.fetchall()
    cursor.close()
    conn.close()
    return render_template('manage_members.html', members=members)


@app.route('/add_member', methods=['GET', 'POST'])
@login_required(roles=['president', 'secretary'])  # Allow secretary to add members
def add_member():
    """Allows President/Secretary to add new members."""
    if request.method == 'POST':
        name = request.form['name']
        username = request.form['username']
        email = request.form['email']
        contact_number = request.form['contact_number']
        pan_number = request.form['pan_number']
        aadhar_number = request.form['aadhar_number']
        password = request.form['password']
        role = request.form.get('role', 'member')  # President/Secretary can specify role for new user

        # Added server-side validation for password
        if not password or not password.strip():  # Ensure password is not empty or just whitespace
            flash('Password cannot be empty.', 'danger')
            return render_template('register.html')

        hashed_password = generate_password_hash(password)

        conn = get_db_connection()
        if conn is None:
            flash('Database connection error.', 'danger')
            return render_template('add_member.html')

        # Use buffered=True for the cursor
        cursor = conn.cursor(buffered=True)
        try:
            cursor.execute(
                "INSERT INTO users (name, username, email, contact_number, pan_number, aadhar_number, password, role) VALUES (%s, %s, %s, %s, %s, %s, %s, %s)",
                (name, username, email, contact_number, pan_number, aadhar_number, hashed_password, role)
            )
            conn.commit()
            flash(f'{name} added successfully as a {role}!', 'success')
            return redirect(url_for('manage_members'))
        except mysql.connector.Error as err:
            if err.errno == 1062:
                flash('Username, Email, PAN, or Aadhar number already exists.', 'danger')
            else:
                flash(f'An error occurred: {err}', 'danger')
            conn.rollback()
        finally:
            cursor.close()
            conn.close()
    return render_template('add_member.html')


@app.route('/edit_member/<int:member_id>', methods=['GET', 'POST'])
@login_required(roles=['president', 'secretary'])  # Allow secretary to edit members
def edit_member(member_id):
    """Allows President/Secretary to edit member details."""
    conn = get_db_connection()
    if conn is None:
        flash('Database connection error. Please try again later.', 'danger')
        return redirect(url_for('manage_members'))
    # Use buffered=True for the cursor
    cursor = conn.cursor(dictionary=True, buffered=True)

    if request.method == 'POST':
        name = request.form['name']
        username = request.form['username']
        email = request.form['email']
        contact_number = request.form['contact_number']
        pan_number = request.form['pan_number']
        aadhar_number = request.form['aadhar_number']
        role = request.form['role']
        password = request.form.get('password')  # Optional password change

        try:
            if password:
                # Only hash and update password if it's provided and not empty
                if not password.strip():  # Check for empty string or just whitespace
                    flash('Password cannot be empty if you choose to change it.', 'danger')
                    # Re-fetch member data to display the form correctly again
                    cursor.execute(
                        "SELECT id, name, username, email, contact_number, pan_number, aadhar_number, role FROM users WHERE id = %s",
                        (member_id,))
                    member = cursor.fetchone()
                    cursor.close()
                    conn.close()
                    return render_template('edit_member.html', member=member)
                hashed_password = generate_password_hash(password)
                cursor.execute(
                    "UPDATE users SET name=%s, username=%s, email=%s, contact_number=%s, pan_number=%s, aadhar_number=%s, password=%s, role=%s WHERE id=%s",
                    (name, username, email, contact_number, pan_number, aadhar_number, hashed_password, role, member_id)
                )
            else:
                cursor.execute(
                    "UPDATE users SET name=%s, username=%s, email=%s, contact_number=%s, pan_number=%s, aadhar_number=%s, role=%s WHERE id=%s",
                    (name, username, email, contact_number, pan_number, aadhar_number, role, member_id)
                )
            conn.commit()
            flash(f'Member {name} updated successfully!', 'success')
            return redirect(url_for('manage_members'))
        except mysql.connector.Error as err:
            if err.errno == 1062:
                flash('Username, Email, PAN, or Aadhar number already exists.', 'danger')
            else:
                flash(f'An error occurred: {err}', 'danger')
            conn.rollback()
        finally:
            cursor.close()
            conn.close()

    # GET request: Display member details for editing
    cursor.execute(
        "SELECT id, name, username, email, contact_number, pan_number, aadhar_number, role FROM users WHERE id = %s",
        (member_id,))
    member = cursor.fetchone()
    cursor.close()  # Close cursor immediately after fetching
    conn.close()
    if not member:
        flash('Member not found.', 'danger')
        return redirect(url_for('manage_members'))
    return render_template('edit_member.html', member=member)


@app.route('/delete_member/<int:member_id>', methods=['POST'])
@login_required(roles=['president', 'secretary'])  # Allow secretary to delete members
def delete_member(member_id):
    """Allows President/Secretary to delete a member."""
    conn = get_db_connection()
    if conn is None:
        flash('Database connection error. Please try again later.', 'danger')
        return redirect(url_for('manage_members'))
    # Use buffered=True for the cursor
    cursor = conn.cursor(buffered=True)
    try:
        cursor.execute("DELETE FROM users WHERE id = %s", (member_id,))
        conn.commit()
        flash('Member deleted successfully!', 'success')
    except mysql.connector.Error as err:
        flash(f'An error occurred: {err}', 'danger')
        conn.rollback()
    finally:
        cursor.close()
        conn.close()
    return redirect(url_for('manage_members'))


@app.route('/member_profile/<int:user_id>')
@login_required()  # Any logged-in user can view their own profile, president can view any
def member_profile(user_id):
    """Displays a member's profile and records."""
    current_user_id = session.get('user_id')
    current_user_role = session.get('role')

    # Allow president or secretary to view any profile, but members can only view their own
    if current_user_role not in ['president', 'secretary'] and current_user_id != user_id:
        flash('You are not authorized to view this profile.', 'danger')
        return redirect(url_for('dashboard'))

    conn = get_db_connection()
    if conn is None:
        flash('Database connection error. Please try again later.', 'danger')
        return redirect(url_for('dashboard'))
    # Use buffered=True for the cursor
    cursor = conn.cursor(dictionary=True, buffered=True)

    # Get user details
    cursor.execute(
        "SELECT id, name, username, email, contact_number, pan_number, aadhar_number, role FROM users WHERE id = %s",
        (user_id,))
    user_profile = cursor.fetchone()
    if not user_profile:
        flash('User not found.', 'danger')
        cursor.close()  # Close cursor before returning
        conn.close()
        return redirect(url_for('dashboard'))

    # Get contributions history
    cursor.execute("SELECT * FROM contributions WHERE user_id = %s ORDER BY year DESC, month DESC", (user_id,))
    contributions = cursor.fetchall()

    # Get loan history
    cursor.execute("SELECT * FROM loans WHERE user_id = %s ORDER BY start_date DESC", (user_id,))
    loans = cursor.fetchall()

    # Get loan payments for each loan
    for loan in loans:
        cursor.execute("SELECT * FROM loan_payments WHERE loan_id = %s ORDER BY payment_date DESC", (loan['id'],))
        loan['payments'] = cursor.fetchall()

    cursor.close()  # Close cursor after all fetches
    conn.close()
    return render_template('member_profile.html', user_profile=user_profile, contributions=contributions, loans=loans)


@app.route('/contributions', methods=['GET', 'POST'])
@login_required()
def contributions():
    """Handles contribution payments and displays history."""
    user_id = session.get('user_id')
    role = session.get('role')
    conn = get_db_connection()
    if conn is None:
        flash('Database connection error. Please try again later.', 'danger')
        return redirect(url_for('dashboard'))
    # Use buffered=True for the cursor
    cursor = conn.cursor(dictionary=True, buffered=True)

    # Get default contribution amount and payment period for display
    default_contribution_amount = Decimal('0.00')
    payment_start_day = 1
    payment_end_day = 7
    default_fine_amount = Decimal('0.00')  # Initialize
    try:
        cursor.execute(
            "SELECT default_contribution_amount, payment_start_day, payment_end_day, default_fine_amount FROM bank_balance WHERE id = 1")
        settings = cursor.fetchone()
        if settings:
            if settings['default_contribution_amount'] is not None:
                default_contribution_amount = settings['default_contribution_amount']
            if settings['payment_start_day'] is not None:
                payment_start_day = settings['payment_start_day']
            if settings['payment_end_day'] is not None:
                payment_end_day = settings['payment_end_day']
            if settings['default_fine_amount'] is not None:
                default_fine_amount = settings['default_fine_amount']
    except Exception as e:
        print(f"Could not fetch default contribution amount or payment period: {e}")

    # Calculate current fine amount based on today's date
    current_day_of_month = datetime.now().day
    current_fine_amount = Decimal('0.00')
    # Fine applies if current day is past the end day, regardless of role
    if current_day_of_month > payment_end_day:
        current_fine_amount = default_fine_amount

    # Calculate total monthly loan interest due for all active loans of the user
    total_monthly_loan_interest_due = Decimal('0.00')
    cursor.execute(
        "SELECT id, amount, interest_rate, start_date FROM loans WHERE user_id = %s AND status IN ('approved', 'overdue')",
        (user_id,))
    active_loans = cursor.fetchall()

    for loan in active_loans:
        cursor.execute(
            "SELECT SUM(amount_paid) as total_paid, SUM(interest_paid) as total_interest_paid_from_payments FROM loan_payments WHERE loan_id = %s",
            (loan['id'],))
        payment_summary = cursor.fetchone()
        total_paid_on_loan = payment_summary['total_paid'] if payment_summary and payment_summary[
            'total_paid'] is not None else Decimal('0.00')
        total_interest_paid_on_loan = payment_summary['total_interest_paid_from_payments'] if payment_summary and \
                                                                                              payment_summary[
                                                                                                  'total_interest_paid_from_payments'] is not None else Decimal(
            '0.00')

        outstanding_principal_on_loan = loan['amount'] - (total_paid_on_loan - total_interest_paid_on_loan)

        # Calculate interest for the current month based on outstanding principal
        if outstanding_principal_on_loan > 0:
            # Determine the start date for interest calculation for this month
            # This is a simplified approach, a real system would need more robust amortization
            # For simplicity, we'll calculate monthly interest on the current outstanding principal.
            monthly_interest_rate_for_loan = loan['interest_rate'] / Decimal('100') / Decimal('12')
            monthly_interest_due_for_this_loan = (
                        outstanding_principal_on_loan * monthly_interest_rate_for_loan).quantize(Decimal('0.01'),
                                                                                                 rounding=ROUND_HALF_UP)
            total_monthly_loan_interest_due += monthly_interest_due_for_this_loan

    # Calculate the total amount to pay (contribution + fine + loan interest)
    total_amount_to_pay = (
                default_contribution_amount + current_fine_amount + total_monthly_loan_interest_due).quantize(
        Decimal('0.01'), rounding=ROUND_HALF_UP)

    if request.method == 'POST':
        # The amount will be taken from the hidden input, which is pre-filled with total_amount_to_pay
        amount_from_form = request.form.get('amount')  # This should be total_amount_to_pay
        utr_number = request.form.get('utr_number', '').strip()

        if not amount_from_form:
            flash('Contribution amount is required.', 'danger')
            cursor.close()
            conn.close()
            return redirect(url_for('contributions'))

        if not utr_number:
            flash('UTR Number is required for contribution.', 'danger')
            cursor.close()
            conn.close()
            return redirect(url_for('contributions'))

        try:
            amount_to_record = Decimal(amount_from_form)  # Use the amount from the form
        except Exception:
            flash('Invalid amount format.', 'danger')
            cursor.close()
            conn.close()
            return redirect(url_for('contributions'))

        current_month = datetime.now().month
        current_year = datetime.now().year
        current_day = datetime.now().day

        # Check if already paid for this month (or pending)
        cursor.execute("SELECT * FROM contributions WHERE user_id = %s AND month = %s AND year = %s",
                       (user_id, current_month, current_year))
        existing_contribution = cursor.fetchone()

        # Recalculate fine_amount for the POST request to ensure consistency
        fine_amount_for_post = Decimal('0.00')
        # Fine applies if current day is past the end day, regardless of role
        if current_day > payment_end_day:
            if not existing_contribution or not existing_contribution['is_paid']:
                fine_amount_for_post = default_fine_amount
                # Flash message for fine will be handled by the initial GET request render or specific POST logic if needed.
                # For now, just ensure the fine is recorded.

        try:
            if existing_contribution and existing_contribution['is_paid']:
                flash(
                    f'You have already paid your contribution for {datetime.strptime(str(current_month), "%m").strftime("%B")} {current_year}.',
                    'info')
                return redirect(url_for('contributions'))

            if existing_contribution:
                cursor.execute(
                    "UPDATE contributions SET amount = %s, utr_number = %s, fine_amount = %s, payment_date = %s WHERE id = %s",
                    (amount_to_record, utr_number, fine_amount_for_post, datetime.now(), existing_contribution['id'])
                )
                flash('Your pending contribution has been updated with the new UTR. Awaiting President approval.',
                      'info')
            else:
                cursor.execute(
                    "INSERT INTO contributions (user_id, amount, month, year, is_paid, fine_amount, utr_number) VALUES (%s, %s, %s, %s, FALSE, %s, %s)",
                    (user_id, amount_to_record, current_month, current_year, fine_amount_for_post, utr_number)
                )
                flash('Contribution submitted for approval. Awaiting President approval.', 'success')

            conn.commit()
            return redirect(url_for('contributions'))
        except mysql.connector.Error as err:
            flash(f'An error occurred while submitting contribution: {err}', 'danger')
            conn.rollback()
        finally:
            cursor.close()
            conn.close()

    # GET request: Display contribution history
    cursor.execute("""
        SELECT c.*, p.name as approver_name
        FROM contributions c
        LEFT JOIN users p ON c.user_id = p.id
        WHERE c.user_id = %s
        ORDER BY c.year DESC, c.month DESC
    """, (user_id,))
    contributions_history = cursor.fetchall()

    # Payment button is always enabled from the UI perspective,
    # but the backend logic still applies fines based on the configured period.
    payment_enabled = True  # Always True now, as per user request (though this variable is no longer used in the template)

    # Check for pending contribution for current month
    current_month = datetime.now().month
    current_year = datetime.now().year
    cursor.execute("""
        SELECT * FROM contributions 
        WHERE user_id = %s AND month = %s AND year = %s AND is_paid = FALSE
    """, (user_id, current_month, current_year))
    pending_contribution_for_display = cursor.fetchone()

    cursor.close()
    conn.close()
    return render_template('contributions.html',
                           contributions_history=contributions_history,
                           payment_enabled=payment_enabled,  # Still passed, but always True
                           pending_contribution=pending_contribution_for_display,  # Pass for display
                           default_contribution_amount=default_contribution_amount,
                           current_fine_amount=current_fine_amount,  # Pass calculated fine
                           total_monthly_loan_interest_due=total_monthly_loan_interest_due,
                           # Pass calculated loan interest
                           total_amount_to_pay=total_amount_to_pay,  # Pass total amount
                           payment_start_day=payment_start_day,  # Pass to template
                           payment_end_day=payment_end_day)  # Pass to template


@app.route('/loans', methods=['GET'])
@login_required()
def loans():
    """Displays loans for members or manages loans for president."""
    user_id = session.get('user_id')
    role = session.get('role')
    conn = get_db_connection()
    if conn is None:
        flash('Database connection error. Please try again later.', 'danger')
        return redirect(url_for('dashboard'))
    # Use buffered=True for the cursor
    cursor = conn.cursor(dictionary=True, buffered=True)

    if role in ['president', 'secretary']:  # Allow secretary to view/manage all loan applications
        # President/Secretary sees all loan applications
        cursor.execute("""
            SELECT l.*, u.name as borrower_name 
            FROM loans l JOIN users u ON l.user_id = u.id 
            ORDER BY l.start_date DESC
        """)
        loans_list = cursor.fetchall()
        template = 'manage_loans.html'
    else:
        # Member sees their own loans
        cursor.execute("""
            SELECT l.*, u.name as president_name 
            FROM loans l LEFT JOIN users u ON l.president_id = u.id 
            WHERE l.user_id = %s ORDER BY l.start_date DESC
        """, (user_id,))
        loans_list = cursor.fetchall()
        template = 'member_loans.html'

    cursor.close()  # Close cursor after all fetches
    conn.close()
    return render_template(template, loans=loans_list)


@app.route('/apply_loan', methods=['GET', 'POST'])
@login_required()
def apply_loan():
    """Allows members to apply for a loan."""
    user_id = session.get('user_id')
    conn = get_db_connection()
    if conn is None:
        flash('Database connection error.', 'danger')
        # Ensure default values are passed even if DB connection fails for initial render
        return render_template('apply_loan.html', default_interest_rate=Decimal('0.00'), bank_balance=Decimal('0.00'),
                               today_date=date.today())

    cursor = conn.cursor(dictionary=True, buffered=True)

    default_interest_rate = Decimal('0.00')  # Initialize with a default
    bank_balance = Decimal('0.00')  # Initialize with a default
    today_date = date.today()  # Get today's date

    try:
        cursor.execute("SELECT default_interest_rate, balance FROM bank_balance WHERE id = 1")
        settings = cursor.fetchone()
        if settings:
            if settings['default_interest_rate'] is not None:
                default_interest_rate = settings['default_interest_rate']
            if settings['balance'] is not None:
                bank_balance = settings['balance']
    except Exception as e:
        print(f"Could not fetch default interest rate or bank balance: {e}")

    if request.method == 'POST':
        # Get form data and strip whitespace
        amount_str = request.form.get('amount', '').strip()
        start_date_str = request.form.get('start_date', '').strip()

        # Input validation for empty strings
        if not amount_str:
            flash('Loan amount is required.', 'danger')
            cursor.close()
            conn.close()
            return render_template('apply_loan.html', default_interest_rate=default_interest_rate,
                                   bank_balance=bank_balance, today_date=today_date)

        if not start_date_str:
            flash('Proposed start date is required.', 'danger')
            cursor.close()
            conn.close()
            return render_template('apply_loan.html', default_interest_rate=default_interest_rate,
                                   bank_balance=bank_balance, today_date=today_date)

        try:
            amount = Decimal(amount_str)
            # Use the default interest rate from settings, not from the disabled form field
            interest_rate = default_interest_rate
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
        except Exception as e:
            flash(f'Invalid data format for loan application: {e}', 'danger')
            cursor.close()
            conn.close()
            return render_template('apply_loan.html', default_interest_rate=default_interest_rate,
                                   bank_balance=bank_balance, today_date=today_date)

        if amount <= 0:
            flash('Loan amount must be positive.', 'danger')
            cursor.close()
            conn.close()
            return render_template('apply_loan.html', default_interest_rate=default_interest_rate,
                                   bank_balance=bank_balance, today_date=today_date)

        if amount > bank_balance:
            flash(f'Requested loan amount (₹{amount:.2f}) exceeds available bank balance (₹{bank_balance:.2f}).',
                  'danger')
            cursor.close()
            conn.close()
            return render_template('apply_loan.html', default_interest_rate=default_interest_rate,
                                   bank_balance=bank_balance, today_date=today_date)

        # Interest rate validation is now based on the fetched default_interest_rate
        if interest_rate < 0 or interest_rate > 100:
            flash('Default interest rate is configured incorrectly (must be between 0 and 100). Please contact admin.',
                  'danger')
            cursor.close()
            conn.close()
            return render_template('apply_loan.html', default_interest_rate=default_interest_rate,
                                   bank_balance=bank_balance, today_date=today_date)

        if start_date < date.today():
            flash('Proposed start date cannot be in the past.', 'danger')
            cursor.close()
            conn.close()
            return render_template('apply_loan.html', default_interest_rate=default_interest_rate,
                                   bank_balance=bank_balance, today_date=today_date)

        try:
            # Loan status starts as 'pending', president_id is NULL until approved
            # The start_date here is the *proposed* start date. It will be updated to the approval date upon approval.
            # end_date is now NULL
            cursor.execute(
                "INSERT INTO loans (user_id, president_id, amount, interest_rate, start_date, status) VALUES (%s, %s, %s, %s, %s, %s)",
                (user_id, None, amount, interest_rate, start_date, 'pending')  # Set president_id to NULL
            )
            conn.commit()
            flash('Loan application submitted successfully! Awaiting President approval.', 'success')
            return redirect(url_for('loans'))
        except mysql.connector.Error as err:
            flash(f'An error occurred while submitting loan application: {err}', 'danger')
            conn.rollback()
        finally:
            cursor.close()
            conn.close()

    # GET request: Display loan application form with default values
    cursor.close()
    conn.close()
    return render_template('apply_loan.html', default_interest_rate=default_interest_rate, bank_balance=bank_balance,
                           today_date=today_date)


@app.route('/approve_loan/<int:loan_id>', methods=['GET'])  # Changed to GET
@login_required(roles=['president', 'secretary'])  # Allow secretary to approve loans
def approve_loan(loan_id):
    """
    Renders the loan disbursement form for the President/Secretary to specify transaction details.
    This route no longer directly approves the loan.
    """
    conn = get_db_connection()
    if conn is None:
        flash('Database connection error. Please try again later.', 'danger')
        return redirect(url_for('loans'))

    cursor = conn.cursor(dictionary=True, buffered=True)
    cursor.execute(
        "SELECT l.id, l.amount, l.user_id, u.name as borrower_name FROM loans l JOIN users u ON l.user_id = u.id WHERE l.id = %s",
        (loan_id,))
    loan = cursor.fetchone()
    cursor.close()
    conn.close()

    if not loan:
        flash('Loan not found.', 'danger')
        return redirect(url_for('loans'))

    # Pass loan details to the new disbursement form
    return render_template('disburse_loan_form.html', loan=loan)


@app.route('/disburse_loan/<int:loan_id>', methods=['POST'])
@login_required(roles=['president', 'secretary'])
def disburse_loan(loan_id):
    """
    Handles the disbursement of an approved loan, recording transaction details.
    This is the new route that performs the actual loan approval and fund deduction.
    """
    conn = get_db_connection()
    if conn is None:
        flash('Database connection error. Please try again later.', 'danger')
        return redirect(url_for('loans'))

    conn.autocommit = False  # Start transaction
    cursor = conn.cursor(buffered=True)  # Cursor is NOT dictionary=True here

    transaction_type = request.form.get('transaction_type')
    disbursement_details = {}

    # Fetch loan amount for validation and deduction
    cursor.execute("SELECT amount FROM loans WHERE id = %s", (loan_id,))
    loan_result = cursor.fetchone()
    if not loan_result:
        flash('Loan not found.', 'danger')
        conn.rollback()
        return redirect(url_for('loans'))
    loan_amount = loan_result[0]  # This is already Decimal

    # Validate transaction details based on type
    if transaction_type == 'cash':
        notes_500_str = request.form.get('notes_500', '0')
        notes_200_str = request.form.get('notes_200', '0')
        notes_100_str = request.form.get('notes_100', '0')

        try:
            notes_500 = int(notes_500_str)
            notes_200 = int(notes_200_str)
            notes_100 = int(notes_100_str)

            if notes_500 < 0 or notes_200 < 0 or notes_100 < 0:
                flash('Note counts cannot be negative.', 'danger')
                conn.rollback()
                return redirect(url_for('approve_loan', loan_id=loan_id))

            calculated_cash_amount = (Decimal(notes_500) * 500) + (Decimal(notes_200) * 200) + (
                        Decimal(notes_100) * 100)

            if calculated_cash_amount != loan_amount:
                flash(
                    f'Cash notes total (₹{calculated_cash_amount:.2f}) does not match loan amount (₹{loan_amount:.2f}).',
                    'danger')
                conn.rollback()
                return redirect(url_for('approve_loan', loan_id=loan_id))

            disbursement_details = {
                'notes_500': notes_500,
                'notes_200': notes_200,
                'notes_100': notes_100
            }
        except ValueError:
            flash('Invalid note count for cash disbursement.', 'danger')
            conn.rollback()
            return redirect(url_for('approve_loan', loan_id=loan_id))

    elif transaction_type == 'cheque':
        cheque_number = request.form.get('cheque_number', '').strip()
        if not cheque_number or not cheque_number.isdigit() or len(cheque_number) != 6:
            flash('Cheque number must be a 6-digit number.', 'danger')
            conn.rollback()
            return redirect(url_for('approve_loan', loan_id=loan_id))
        disbursement_details = {'cheque_number': cheque_number}

    elif transaction_type == 'upi':
        upi_utr = request.form.get('upi_utr', '').strip()
        if not upi_utr or not upi_utr.isdigit() or len(upi_utr) != 12:
            flash('UPI UTR must be a 12-digit number.', 'danger')
            conn.rollback()
            return redirect(url_for('approve_loan', loan_id=loan_id))
        disbursement_details = {'upi_utr': upi_utr}
    else:
        flash('Invalid transaction type selected.', 'danger')
        conn.rollback()
        return redirect(url_for('approve_loan', loan_id=loan_id))

    # Convert disbursement_details to JSON string for storage
    disbursement_details_json = json.dumps(disbursement_details)

    try:
        # 1. Get current bank balance
        cursor.execute("SELECT balance FROM bank_balance WHERE id = 1")
        balance_result = cursor.fetchone()  # This will be a tuple (balance,)
        current_balance = balance_result[0] if balance_result and balance_result[0] is not None else Decimal('0.00')

        # 2. Check for sufficient funds
        if current_balance < loan_amount:
            flash(
                f'Insufficient bank balance (Current: ₹{current_balance:.2f}, Required: ₹{loan_amount:.2f}) to approve this loan. Please deposit funds.',
                'danger')
            conn.rollback()
            return redirect(url_for('loans'))

        # 3. Update loan status AND set start_date to today's date (approval date)
        # Also store disbursement type and details
        cursor.execute(
            "UPDATE loans SET status = 'approved', president_id = %s, start_date = %s, disbursement_type = %s, disbursement_details = %s WHERE id = %s",
            (session['user_id'], date.today(), transaction_type, disbursement_details_json, loan_id)
        )

        # 4. Deduct loan amount from bank balance
        cursor.execute("UPDATE bank_balance SET balance = balance - %s WHERE id = 1", (loan_amount,))

        # 5. Commit the transaction
        conn.commit()
        flash('Loan approved and amount disbursed! Transaction details recorded.', 'success')

    except mysql.connector.Error as err:
        flash(f'An error occurred while disbursing loan: {err}', 'danger')
        conn.rollback()
    finally:
        cursor.close()
        conn.close()
    return redirect(url_for('loans'))


@app.route('/reject_loan/<int:loan_id>', methods=['POST'])
@login_required(roles=['president', 'secretary'])  # Allow secretary to reject loans
def reject_loan(loan_id):
    """Allows President/Secretary to reject a loan."""
    conn = get_db_connection()
    if conn is None:
        flash('Database connection error. Please try again later.', 'danger')
        return redirect(url_for('loans'))
    # Use buffered=True for the cursor
    cursor = conn.cursor(buffered=True)
    try:
        # Record which president/secretary rejected the loan
        cursor.execute("UPDATE loans SET status = 'rejected', president_id = %s WHERE id = %s",
                       (session['user_id'], loan_id))
        conn.commit()
        flash('Loan rejected.', 'info')
    except mysql.connector.Error as err:
        flash(f'An error occurred while rejecting loan: {err}', 'danger')
        conn.rollback()
    finally:
        cursor.close()
        conn.close()
    return redirect(url_for('loans'))


@app.route('/record_loan_payment/<int:loan_id>', methods=['GET', 'POST'])
@login_required()  # Can be done by member or treasurer/president/secretary
def record_loan_payment(loan_id):
    """Allows recording a payment for a specific loan."""
    conn = get_db_connection()
    if conn is None:
        flash('Database connection error. Please try again later.', 'danger')
        return redirect(url_for('loans'))
    # Use buffered=True for the cursor
    cursor = conn.cursor(dictionary=True, buffered=True)

    # Get loan details
    cursor.execute("SELECT id, user_id, amount, interest_rate, start_date, status FROM loans WHERE id = %s", (loan_id,))
    loan = cursor.fetchone()

    if not loan:
        flash('Loan not found.', 'danger')
        cursor.close()
        conn.close()
        return redirect(url_for('loans'))

    # Ensure only the borrower or president/secretary can record payments
    if session.get('user_id') != loan['user_id'] and session.get('role') not in ['president', 'secretary']:
        flash('You are not authorized to record payments for this loan.', 'danger')
        cursor.close()
        conn.close()
        return redirect(url_for('loans'))

    if loan['status'] == 'completed':
        flash('This loan is already completed. No more payments are needed.', 'info')
        cursor.close()
        conn.close()
        return redirect(url_for('loans'))

    # Calculate current outstanding principal and total interest paid so far
    cursor.execute(
        "SELECT SUM(amount_paid) as total_paid, SUM(interest_paid) as total_interest_paid_from_payments FROM loan_payments WHERE loan_id = %s",
        (loan_id,))
    payment_summary = cursor.fetchone()
    total_paid_so_far = payment_summary['total_paid'] if payment_summary and payment_summary[
        'total_paid'] is not None else Decimal('0.00')
    total_interest_paid_from_payments = payment_summary['total_interest_paid_from_payments'] if payment_summary and \
                                                                                                payment_summary[
                                                                                                    'total_interest_paid_from_payments'] is not None else Decimal(
        '0.00')

    # Initial principal is the loan amount
    outstanding_principal = loan['amount'] - (total_paid_so_far - total_interest_paid_from_payments)

    # Calculate monthly interest based on outstanding principal (simple interest for demo)
    # Assuming interest_rate is annual percentage
    monthly_interest_rate = loan['interest_rate'] / Decimal('100') / Decimal('12')
    monthly_interest_due = (outstanding_principal * monthly_interest_rate).quantize(Decimal('0.01'),
                                                                                    rounding=ROUND_HALF_UP)

    if request.method == 'POST':
        amount_paid_str = request.form.get('amount_paid')

        if not amount_paid_str:
            flash('Payment amount is required.', 'danger')
            cursor.close()
            conn.close()
            return redirect(url_for('record_loan_payment', loan_id=loan_id))

        try:
            amount_paid = Decimal(amount_paid_str)
        except Exception:
            flash('Invalid amount format.', 'danger')
            cursor.close()
            conn.close()
            return redirect(url_for('record_loan_payment', loan_id=loan_id))

        if amount_paid <= 0:
            flash('Payment amount must be positive.', 'danger')
            cursor.close()
            conn.close()
            return redirect(url_for('record_loan_payment', loan_id=loan_id))

        # Determine how much goes to interest and how much to principal
        interest_portion = min(amount_paid, monthly_interest_due)
        principal_portion = amount_paid - interest_portion

        try:
            cursor.execute(
                "INSERT INTO loan_payments (loan_id, amount_paid, interest_paid, payment_date) VALUES (%s, %s, %s, %s)",
                (loan_id, amount_paid, interest_portion, datetime.now())
            )

            # Update bank balance (money coming back to the gat)
            cursor.execute("UPDATE bank_balance SET balance = balance + %s WHERE id = 1", (amount_paid,))

            # Recalculate outstanding principal after this payment
            new_outstanding_principal = outstanding_principal - principal_portion

            # Check if loan is fully paid (principal is zero or less)
            if new_outstanding_principal <= Decimal('0.00'):
                cursor.execute("UPDATE loans SET status = 'completed', actual_end_date = %s WHERE id = %s",
                               (date.today(), loan_id))
                flash('Loan fully paid and marked as completed!', 'success')
            else:
                flash('Loan payment recorded successfully!', 'success')

            conn.commit()
            return redirect(url_for('loans'))
        except mysql.connector.Error as err:
            flash(f'An error occurred while recording payment: {err}', 'danger')
            conn.rollback()
        finally:
            cursor.close()
            conn.close()

    # GET request: Display loan payment form
    # Recalculate for display after potential POST redirect or initial GET
    cursor.execute("SELECT id, user_id, amount, interest_rate, start_date, status FROM loans WHERE id = %s", (loan_id,))
    loan = cursor.fetchone()  # Re-fetch to ensure latest state

    cursor.execute(
        "SELECT SUM(amount_paid) as total_paid, SUM(interest_paid) as total_interest_paid_from_payments FROM loan_payments WHERE loan_id = %s",
        (loan_id,))
    payment_summary = cursor.fetchone()
    total_paid_so_far = payment_summary['total_paid'] if payment_summary and payment_summary[
        'total_paid'] is not None else Decimal('0.00')
    total_interest_paid_from_payments = payment_summary['total_interest_paid_from_payments'] if payment_summary and \
                                                                                                payment_summary[
                                                                                                    'total_interest_paid_from_payments'] is not None else Decimal(
        '0.00')

    outstanding_principal = loan['amount'] - (total_paid_so_far - total_interest_paid_from_payments)
    monthly_interest_rate = loan['interest_rate'] / Decimal('100') / Decimal('12')
    monthly_interest_due = (outstanding_principal * monthly_interest_rate).quantize(Decimal('0.01'),
                                                                                    rounding=ROUND_HALF_UP)

    # total_expected_repayment_over_term and loan_duration_months removed as end_date is no longer used.

    remaining_total_amount = (outstanding_principal + monthly_interest_due).quantize(Decimal('0.01'),
                                                                                     rounding=ROUND_HALF_UP)
    if remaining_total_amount < 0:
        remaining_total_amount = Decimal('0.00')

    cursor.close()
    conn.close()
    return render_template('record_loan_payment.html',
                           loan=loan,
                           total_paid=total_paid_so_far,
                           outstanding_principal=outstanding_principal,
                           monthly_interest_due=monthly_interest_due,
                           remaining_total_amount=remaining_total_amount,
                           total_interest_paid_from_payments=total_interest_paid_from_payments
                           )


# New route for President/Secretary to manage settings (e.g., default fine amount and interest rate)
@app.route('/manage_settings', methods=['GET', 'POST'])
@login_required(roles=['president', 'secretary'])
def manage_settings():
    """Allows President/Secretary to manage application settings like default fine amount, interest rate, contribution amount, and payment period."""
    conn = get_db_connection()
    if conn is None:
        flash('Database connection error. Please try again later.', 'danger')
        return redirect(url_for('dashboard'))
    cursor = conn.cursor(dictionary=True, buffered=True)

    if request.method == 'POST':
        default_fine_amount_str = request.form.get('default_fine_amount')
        default_interest_rate_str = request.form.get('default_interest_rate')
        default_contribution_amount_str = request.form.get('default_contribution_amount')
        payment_start_day_str = request.form.get('payment_start_day')  # New field
        payment_end_day_str = request.form.get('payment_end_day')  # New field

        if not all([default_fine_amount_str, default_interest_rate_str, default_contribution_amount_str,
                    payment_start_day_str, payment_end_day_str]):
            flash('All default settings (fine, interest, contribution amounts, and payment period) are required.',
                  'danger')
            # Re-fetch current settings to pre-fill the form on error
            cursor.execute(
                "SELECT default_fine_amount, default_interest_rate, default_contribution_amount, payment_start_day, payment_end_day FROM bank_balance WHERE id = 1")
            settings = cursor.fetchone()
            current_fine_amount = settings['default_fine_amount'] if settings and settings[
                'default_fine_amount'] is not None else Decimal('0.00')
            current_interest_rate = settings['default_interest_rate'] if settings and settings[
                'default_interest_rate'] is not None else Decimal('0.00')
            current_contribution_amount = settings['default_contribution_amount'] if settings and settings[
                'default_contribution_amount'] is not None else Decimal('0.00')
            current_payment_start_day = settings['payment_start_day'] if settings and settings[
                'payment_start_day'] is not None else 1
            current_payment_end_day = settings['payment_end_day'] if settings and settings[
                'payment_end_day'] is not None else 7
            cursor.close()
            conn.close()
            return render_template('manage_settings.html',
                                   current_fine_amount=current_fine_amount,
                                   current_interest_rate=current_interest_rate,
                                   current_contribution_amount=current_contribution_amount,
                                   current_payment_start_day=current_payment_start_day,
                                   current_payment_end_day=current_payment_end_day)

        try:
            new_fine_amount = Decimal(default_fine_amount_str)
            new_interest_rate = Decimal(default_interest_rate_str)
            new_contribution_amount = Decimal(default_contribution_amount_str)
            new_payment_start_day = int(payment_start_day_str)
            new_payment_end_day = int(payment_end_day_str)

            if new_fine_amount < 0:
                flash('Fine amount cannot be negative.', 'danger')
                cursor.close()
                conn.close()
                return render_template('manage_settings.html',
                                       current_fine_amount=new_fine_amount,
                                       current_interest_rate=new_interest_rate,
                                       current_contribution_amount=new_contribution_amount,
                                       current_payment_start_day=new_payment_start_day,
                                       current_payment_end_day=new_payment_end_day)
            if new_interest_rate < 0 or new_interest_rate > 100:
                flash('Interest rate must be between 0 and 100.', 'danger')
                cursor.close()
                conn.close()
                return render_template('manage_settings.html',
                                       current_fine_amount=new_fine_amount,
                                       current_interest_rate=new_interest_rate,
                                       current_contribution_amount=new_contribution_amount,
                                       current_payment_start_day=new_payment_start_day,
                                       current_payment_end_day=new_payment_end_day)
            if new_contribution_amount <= 0:
                flash('Contribution amount must be positive.', 'danger')
                cursor.close()
                conn.close()
                return render_template('manage_settings.html',
                                       current_fine_amount=new_fine_amount,
                                       current_interest_rate=new_interest_rate,
                                       current_contribution_amount=new_contribution_amount,
                                       current_payment_start_day=new_payment_start_day,
                                       current_payment_end_day=new_payment_end_day)

            if not (1 <= new_payment_start_day <= 31) or not (1 <= new_payment_end_day <= 31):
                flash('Payment start and end days must be between 1 and 31.', 'danger')
                cursor.close()
                conn.close()
                return render_template('manage_settings.html',
                                       current_fine_amount=new_fine_amount,
                                       current_interest_rate=new_interest_rate,
                                       current_contribution_amount=new_contribution_amount,
                                       current_payment_start_day=new_payment_start_day,
                                       current_payment_end_day=new_payment_end_day)

            if new_payment_start_day > new_payment_end_day:
                flash('Payment start day cannot be after the end day.', 'danger')
                cursor.close()
                conn.close()
                return render_template('manage_settings.html',
                                       current_fine_amount=new_fine_amount,
                                       current_interest_rate=new_interest_rate,
                                       current_contribution_amount=new_contribution_amount,
                                       current_payment_start_day=new_payment_start_day,
                                       current_payment_end_day=new_payment_end_day)


        except Exception:
            flash('Invalid number format for fine, interest, contribution amount, or payment days.', 'danger')
            cursor.close()
            conn.close()
            return render_template('manage_settings.html',
                                   current_fine_amount=Decimal('0.00'),
                                   current_interest_rate=Decimal('0.00'),
                                   current_contribution_amount=Decimal('0.00'),
                                   current_payment_start_day=1,
                                   current_payment_end_day=7)

        try:
            # Assuming bank_balance table has a single row with id = 1 for settings
            cursor.execute(
                "UPDATE bank_balance SET default_fine_amount = %s, default_interest_rate = %s, default_contribution_amount = %s, payment_start_day = %s, payment_end_day = %s WHERE id = 1",
                (new_fine_amount, new_interest_rate, new_contribution_amount, new_payment_start_day,
                 new_payment_end_day))
            conn.commit()
            flash(
                f'Settings updated successfully! Default Fine: ₹{new_fine_amount:.2f}, Default Interest: {new_interest_rate:.2f}%, Default Contribution: ₹{new_contribution_amount:.2f}, Payment Period: {new_payment_start_day} to {new_payment_end_day}.',
                'success')
        except mysql.connector.Error as err:
            flash(f'An error occurred while updating settings: {err}', 'danger')
            conn.rollback()
        finally:
            cursor.close()
            conn.close()
        return redirect(url_for('manage_settings'))

    # GET request: Display current settings
    cursor.execute(
        "SELECT default_fine_amount, default_interest_rate, default_contribution_amount, payment_start_day, payment_end_day FROM bank_balance WHERE id = 1")
    settings = cursor.fetchone()
    current_fine_amount = settings['default_fine_amount'] if settings and settings[
        'default_fine_amount'] is not None else Decimal('0.00')
    current_interest_rate = settings['default_interest_rate'] if settings and settings[
        'default_interest_rate'] is not None else Decimal('0.00')
    current_contribution_amount = settings['default_contribution_amount'] if settings and settings[
        'default_contribution_amount'] is not None else Decimal('0.00')
    current_payment_start_day = settings['payment_start_day'] if settings and settings[
        'payment_start_day'] is not None else 1
    current_payment_end_day = settings['payment_end_day'] if settings and settings['payment_end_day'] is not None else 7

    cursor.close()
    conn.close()
    return render_template('manage_settings.html',
                           current_fine_amount=current_fine_amount,
                           current_interest_rate=current_interest_rate,
                           current_contribution_amount=current_contribution_amount,
                           current_payment_start_day=current_payment_start_day,
                           current_payment_end_day=current_payment_end_day)


@app.route('/bank_balance', methods=['GET', 'POST'])
@login_required(roles=['president', 'secretary'])  # Allow secretary to manage bank balance
def bank_balance():
    """Manages and displays the collective bank balance."""
    conn = get_db_connection()
    if conn is None:
        flash('Database connection error. Please try again later.', 'danger')
        return redirect(url_for('dashboard'))
    # Use buffered=True for the cursor
    cursor = conn.cursor(dictionary=True, buffered=True)

    if request.method == 'POST':
        action = request.form['action']
        amount_str = request.form.get('amount')

        if not amount_str:
            flash('Amount is required.', 'danger')
            return redirect(url_for('bank_balance'))

        try:
            amount = Decimal(amount_str)  # Convert to Decimal
        except Exception:
            flash('Invalid amount format.', 'danger')
            return redirect(url_for('bank_balance'))

        if amount <= 0:
            flash('Amount must be positive.', 'danger')
            return redirect(url_for('bank_balance'))

        try:
            cursor.execute("SELECT balance FROM bank_balance WHERE id = 1")
            result_balance = cursor.fetchone()
            current_balance = result_balance[0] if result_balance and result_balance[0] is not None else Decimal('0.00')

            if action == 'deposit':
                new_balance = current_balance + amount
                flash(f'Deposited {amount:.2f} into bank balance.', 'success')
            elif action == 'withdraw':
                if current_balance < amount:  # Compare Decimal with Decimal
                    flash('Insufficient balance for withdrawal.', 'danger')
                    conn.rollback()
                    cursor.close()  # Close cursor before redirect
                    conn.close()
                    return redirect(url_for('bank_balance'))
                new_balance = current_balance - amount
                flash(f'Withdrew {amount:.2f} from bank balance.', 'success')
            else:
                flash('Invalid action.', 'danger')
                conn.rollback()
                cursor.close()  # Close cursor before redirect
                conn.close()
                return redirect(url_for('bank_balance'))

            cursor.execute("UPDATE bank_balance SET balance = %s WHERE id = 1", (new_balance,))
            conn.commit()
        except mysql.connector.Error as err:
            flash(f'An error occurred: {err}', 'danger')
            conn.rollback()
        finally:
            cursor.close()  # Close cursor after all operations
            conn.close()
        return redirect(url_for('bank_balance'))

    # GET request: Display current balance
    cursor.execute("SELECT balance, last_updated FROM bank_balance WHERE id = 1")
    balance_info = cursor.fetchone()
    cursor.close()  # Close cursor after all fetches
    conn.close()
    return render_template('bank_balance.html', balance_info=balance_info)


@app.route('/send_reminders', methods=['POST'])
@login_required(roles=['president', 'secretary'])  # Allow secretary to send reminders
def send_reminders():
    """Simulates sending reminders for contributions and interest."""
    conn = get_db_connection()
    if conn is None:
        flash('Database connection error. Please try again later.', 'danger')
        return redirect(url_for('dashboard'))
    # Use buffered=True for the cursor
    cursor = conn.cursor(dictionary=True, buffered=True)

    current_month = datetime.now().month
    current_year = datetime.now().year

    # Find members who haven't paid contribution for current month
    cursor.execute("""
        SELECT u.name, u.email, u.contact_number 
        FROM users u 
        LEFT JOIN contributions c ON u.id = c.user_id AND c.month = %s AND c.year = %s AND c.is_paid = TRUE
        WHERE u.role = 'member' AND c.id IS NULL
    """, (current_month, current_year))
    unpaid_members = cursor.fetchall()

    # Find members with active loans
    cursor.execute("""
        SELECT DISTINCT u.name, u.email, u.contact_number 
        FROM users u 
        JOIN loans l ON u.id = l.user_id 
        WHERE l.status IN ('approved', 'overdue')
    """)
    loan_members = cursor.fetchall()

    # In a real application, you would integrate with SMS/Email APIs here.
    # For this example, we'll just flash a message.
    if unpaid_members:
        flash(f'Reminders sent to {len(unpaid_members)} members for pending contributions.', 'info')
        # print(f"Contribution reminders to: {unpaid_members}") # For debugging
    if loan_members:
        flash(f'Reminders sent to {len(loan_members)} members for loan interest/payments.', 'info')
        # print(f"Loan reminders to: {loan_members}") # For debugging

    if not unpaid_members and not loan_members:
        flash('No pending contributions or active loans found for reminders.', 'info')

    cursor.close()  # Close cursor after all fetches
    conn.close()
    return redirect(url_for('dashboard'))  # Redirect back to the main dashboard


# New route for President/Secretary to review and edit pending loan applications
@app.route('/review_loan/<int:loan_id>', methods=['GET', 'POST'])
@login_required(roles=['president', 'secretary'])
def review_loan(loan_id):
    """Allows President/Secretary to review and edit a pending loan application."""
    conn = get_db_connection()
    if conn is None:
        flash('Database connection error. Please try again later.', 'danger')
        return redirect(url_for('loans'))
    cursor = conn.cursor(dictionary=True, buffered=True)

    if request.method == 'POST':
        amount_str = request.form.get('amount')
        interest_rate_str = request.form.get('interest_rate')
        start_date_str = request.form.get('start_date')
        # end_date_str removed from form

        # Input validation
        if not all([amount_str, interest_rate_str, start_date_str]):  # Removed end_date_str
            flash('All loan fields are required.', 'danger')
            # Re-fetch loan details to re-render the form with current values
            cursor.execute(
                "SELECT l.*, u.name as borrower_name FROM loans l JOIN users u ON l.user_id = u.id WHERE l.id = %s",
                (loan_id,))
            loan_details = cursor.fetchone()
            cursor.close()
            conn.close()
            return render_template('review_loan.html', loan=loan_details)

        try:
            amount = Decimal(amount_str)
            interest_rate = Decimal(interest_rate_str)
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
            # end_date removed
        except Exception as e:
            flash(f'Invalid data format: {e}', 'danger')
            cursor.execute(
                "SELECT l.*, u.name as borrower_name FROM loans l JOIN users u ON l.user_id = u.id WHERE l.id = %s",
                (loan_id,))
            loan_details = cursor.fetchone()
            cursor.close()
            conn.close()
            return render_template('review_loan.html', loan=loan_details)

        if amount <= 0 or interest_rate <= 0:
            flash('Amount and interest rate must be positive values.', 'danger')
            cursor.execute(
                "SELECT l.*, u.name as borrower_name FROM loans l JOIN users u ON l.user_id = u.id WHERE l.id = %s",
                (loan_id,))
            loan_details = cursor.fetchone()
            cursor.close()
            conn.close()
            return render_template('review_loan.html', loan=loan_details)

        # No end_date validation needed

        try:
            # Update statement no longer includes end_date
            cursor.execute(
                "UPDATE loans SET amount = %s, interest_rate = %s, start_date = %s WHERE id = %s",
                (amount, interest_rate, start_date, loan_id)
            )
            conn.commit()
            flash('Loan application updated successfully! You can now approve or reject it.', 'success')
            return redirect(url_for('loans'))  # Redirect back to manage loans
        except mysql.connector.Error as err:
            flash(f'An error occurred while updating loan: {err}', 'danger')
            conn.rollback()
        finally:
            cursor.close()
            conn.close()

    # GET request: Display current loan details for review
    cursor.execute("SELECT l.*, u.name as borrower_name FROM loans l JOIN users u ON l.user_id = u.id WHERE l.id = %s",
                   (loan_id,))
    loan_details = cursor.fetchone()
    cursor.close()
    conn.close()

    if not loan_details:
        flash('Loan application not found.', 'danger')
        return redirect(url_for('loans'))

    return render_template('review_loan.html', loan=loan_details)


@app.route('/close_loan/<int:loan_id>', methods=['GET', 'POST'])
@login_required()
def close_loan(loan_id):
    """Allows a member to apply to close their loan."""
    user_id = session.get('user_id')
    role = session.get('role')
    conn = get_db_connection()
    if conn is None:
        flash('Database connection error. Please try again later.', 'danger')
        return redirect(url_for('loans'))
    cursor = conn.cursor(dictionary=True, buffered=True)

    # Get loan details
    cursor.execute("SELECT id, user_id, amount, interest_rate, start_date, status FROM loans WHERE id = %s", (loan_id,))
    loan = cursor.fetchone()

    if not loan:
        flash('Loan not found.', 'danger')
        cursor.close()
        conn.close()
        return redirect(url_for('loans'))

    # Ensure only the borrower or president/secretary can close this loan
    if user_id != loan['user_id'] and role not in ['president', 'secretary']:
        flash('You are not authorized to close this loan.', 'danger')
        cursor.close()
        conn.close()
        return redirect(url_for('loans'))

    if loan['status'] == 'completed':
        flash('This loan is already completed. No more payments are needed.', 'info')
        cursor.close()
        conn.close()
        return redirect(url_for('loans'))

    # Calculate current outstanding principal and total interest paid so far
    cursor.execute(
        "SELECT SUM(amount_paid) as total_paid, SUM(interest_paid) as total_interest_paid_from_payments FROM loan_payments WHERE loan_id = %s",
        (loan_id,))
    payment_summary = cursor.fetchone()
    total_paid_so_far = payment_summary['total_paid'] if payment_summary and payment_summary[
        'total_paid'] is not None else Decimal('0.00')
    total_interest_paid_from_payments = payment_summary['total_interest_paid_from_payments'] if payment_summary and \
                                                                                                payment_summary[
                                                                                                    'total_interest_paid_from_payments'] is not None else Decimal(
        '0.00')

    outstanding_principal = loan['amount'] - (total_paid_so_far - total_interest_paid_from_payments)

    # Calculate accrued interest since last payment or loan start
    last_payment_date = None
    cursor.execute("SELECT MAX(payment_date) as last_payment_date FROM loan_payments WHERE loan_id = %s", (loan_id,))
    last_payment_result = cursor.fetchone()
    if last_payment_result and last_payment_result['last_payment_date']:
        last_payment_date = last_payment_result['last_payment_date'].date()  # Convert datetime to date

    interest_start_date = last_payment_date if last_payment_date else loan['start_date']

    # Ensure interest calculation is only for active loans and within loan term
    if loan['status'] in ['approved', 'overdue']:
        today = date.today()
        # Calculate days since last interest calculation or loan start
        days_since_last_calc = (today - interest_start_date).days

        # Simple daily interest calculation
        daily_interest_rate = loan['interest_rate'] / Decimal('100') / Decimal('365')
        accrued_interest = (outstanding_principal * daily_interest_rate * days_since_last_calc).quantize(
            Decimal('0.01'), rounding=ROUND_HALF_UP)
    else:
        accrued_interest = Decimal('0.00')  # No new interest if loan is not active

    remaining_amount_to_close = (outstanding_principal + accrued_interest).quantize(Decimal('0.01'),
                                                                                    rounding=ROUND_HALF_UP)
    if remaining_amount_to_close < 0:
        remaining_amount_to_close = Decimal('0.00')

    if request.method == 'POST':
        closing_amount_str = request.form.get('closing_amount')

        if not closing_amount_str:
            flash('Closing amount is required.', 'danger')
            cursor.close()
            conn.close()
            return render_template('close_loan.html', loan=loan,
                                   outstanding_principal=outstanding_principal,
                                   accrued_interest=accrued_interest,
                                   remaining_amount_to_close=remaining_amount_to_close)
        try:
            closing_amount = Decimal(closing_amount_str)
        except Exception:
            flash('Invalid amount format.', 'danger')
            cursor.close()
            conn.close()
            return render_template('close_loan.html', loan=loan,
                                   outstanding_principal=outstanding_principal,
                                   accrued_interest=accrued_interest,
                                   remaining_amount_to_close=remaining_amount_to_close)

        if closing_amount <= 0:
            flash('Closing amount must be positive.', 'danger')
            cursor.close()
            conn.close()
            return render_template('close_loan.html', loan=loan,
                                   outstanding_principal=outstanding_principal,
                                   accrued_interest=accrued_interest,
                                   remaining_amount_to_close=remaining_amount_to_close)

        if closing_amount < remaining_amount_to_close:
            flash(
                f'The closing amount is less than the total outstanding amount (Principal + Accrued Interest: ₹{remaining_amount_to_close:.2f}). Please pay the full amount to close the loan.',
                'danger')
            cursor.close()
            conn.close()
            return render_template('close_loan.html', loan=loan,
                                   outstanding_principal=outstanding_principal,
                                   accrued_interest=accrued_interest,
                                   remaining_amount_to_close=remaining_amount_to_close)

        try:
            # Record the final payment
            cursor.execute(
                "INSERT INTO loan_payments (loan_id, amount_paid, interest_paid, payment_date) VALUES (%s, %s, %s, %s)",
                (loan_id, closing_amount, accrued_interest, datetime.now())
            )

            # Update bank balance
            cursor.execute("UPDATE bank_balance SET balance = balance + %s WHERE id = 1", (closing_amount,))

            # Mark loan as completed and set actual_end_date
            cursor.execute("UPDATE loans SET status = 'completed', actual_end_date = %s WHERE id = %s",
                           (date.today(), loan_id))

            conn.commit()
            flash('Loan successfully closed!', 'success')
            return redirect(url_for('loans'))
        except mysql.connector.Error as err:
            flash(f'An error occurred while closing loan: {err}', 'danger')
            conn.rollback()
        finally:
            cursor.close()
            conn.close()

    cursor.close()
    conn.close()
    return render_template('close_loan.html',
                           loan=loan,
                           outstanding_principal=outstanding_principal,
                           accrued_interest=accrued_interest,
                           remaining_amount_to_close=remaining_amount_to_close)


# New routes for Contribution Approval Workflow
@app.route('/manage_contributions', methods=['GET'])
@login_required(roles=['president', 'secretary'])
def manage_contributions():
    """Allows President/Secretary to view and manage pending contributions."""
    conn = get_db_connection()
    if conn is None:
        flash('Database connection error. Please try again later.', 'danger')
        return redirect(url_for('dashboard'))
    cursor = conn.cursor(dictionary=True, buffered=True)

    # Fetch all pending contributions (is_paid = FALSE and president_utr_number is NULL)
    cursor.execute("""
        SELECT c.*, u.name as member_name, u.username as member_username
        FROM contributions c
        JOIN users u ON c.user_id = u.id
        WHERE c.is_paid = FALSE AND c.president_utr_number IS NULL
        ORDER BY c.payment_date DESC
    """)
    pending_contributions = cursor.fetchall()

    # Fetch all approved contributions for this month
    current_month = datetime.now().month
    current_year = datetime.now().year
    cursor.execute("""
        SELECT c.*, u.name as member_name, u.username as member_username, p.name as approver_name
        FROM contributions c
        JOIN users u ON c.user_id = u.id
        LEFT JOIN users p ON c.president_id = p.id
        WHERE c.is_paid = TRUE AND c.month = %s AND c.year = %s
        ORDER BY c.payment_date DESC
    """, (current_month, current_year))
    approved_contributions_this_month = cursor.fetchall()

    cursor.close()
    conn.close()
    return render_template('manage_contributions.html',
                           pending_contributions=pending_contributions,
                           approved_contributions_this_month=approved_contributions_this_month)


@app.route('/approve_contribution/<int:contribution_id>', methods=['POST'])
@login_required(roles=['president', 'secretary'])
def approve_contribution(contribution_id):
    """Allows President/Secretary to approve a contribution after UTR matching."""
    president_utr_number = request.form.get('president_utr_number', '').strip()

    if not president_utr_number:
        flash('President UTR Number is required to approve this contribution.', 'danger')
        return redirect(url_for('manage_contributions'))

    conn = get_db_connection()
    if conn is None:
        flash('Database connection error. Please try again later.', 'danger')
        return redirect(url_for('manage_contributions'))

    conn.autocommit = False  # Start transaction

    cursor = conn.cursor(dictionary=True, buffered=True)

    try:
        # 1. Fetch the pending contribution
        cursor.execute("SELECT id, user_id, amount, fine_amount, utr_number, is_paid FROM contributions WHERE id = %s",
                       (contribution_id,))
        contribution = cursor.fetchone()

        if not contribution:
            flash('Contribution not found.', 'danger')
            conn.rollback()
            return redirect(url_for('manage_contributions'))

        if contribution['is_paid']:  # Already paid/approved
            flash('This contribution has already been approved.', 'info')
            conn.rollback()
            return redirect(url_for('manage_contributions'))

        # 2. Compare UTR numbers
        if contribution['utr_number'] != president_utr_number:
            flash(f'UTR Number mismatch for contribution ID {contribution_id}. Please check the UTR entered.', 'danger')
            conn.rollback()
            return redirect(url_for('manage_contributions'))

        # 3. Update contribution status to paid and record approver
        total_amount_to_add = contribution['amount'] + contribution['fine_amount']

        cursor.execute(
            "UPDATE contributions SET is_paid = TRUE, president_id = %s, president_utr_number = %s, payment_date = %s WHERE id = %s",
            (session['user_id'], president_utr_number, datetime.now(), contribution_id)
        )

        # 4. Update bank balance
        cursor.execute("UPDATE bank_balance SET balance = balance + %s WHERE id = 1", (total_amount_to_add,))

        conn.commit()
        flash(
            f'Contribution from {contribution["user_id"]} approved successfully! Amount ₹{total_amount_to_add:.2f} added to bank balance.',
            'success')

    except mysql.connector.Error as err:
        flash(f'An error occurred while approving contribution: {err}', 'danger')
        conn.rollback()
    finally:
        cursor.close()
        conn.close()
    return redirect(url_for('manage_contributions'))


@app.route('/reject_contribution/<int:contribution_id>', methods=['POST'])
@login_required(roles=['president', 'secretary'])
def reject_contribution(contribution_id):
    """Allows President/Secretary to reject a contribution."""
    conn = get_db_connection()
    if conn is None:
        flash('Database connection error. Please try again later.', 'danger')
        return redirect(url_for('manage_contributions'))

    cursor = conn.cursor(buffered=True)

    try:
        # Fetch the contribution to ensure it's pending
        cursor.execute("SELECT is_paid FROM contributions WHERE id = %s", (contribution_id,))
        contribution_status = cursor.fetchone()

        if not contribution_status:
            flash('Contribution not found.', 'danger')
            return redirect(url_for('manage_contributions'))

        if contribution_status[0]:  # is_paid is True
            flash('This contribution has already been approved and cannot be rejected.', 'info')
            return redirect(url_for('manage_contributions'))

        # Mark as rejected (by setting president_utr_number to a specific string or adding a 'status' column)
        # For simplicity, we'll just set president_id to mark who rejected, and keep is_paid=FALSE
        cursor.execute(
            "UPDATE contributions SET president_id = %s, president_utr_number = 'REJECTED' WHERE id = %s",
            (session['user_id'], contribution_id)
        )
        conn.commit()
        flash('Contribution rejected.', 'info')

    except mysql.connector.Error as err:
        flash(f'An error occurred while rejecting contribution: {err}', 'danger')
        conn.rollback()
    finally:
        cursor.close()
        conn.close()
    return redirect(url_for('manage_contributions'))


@app.route('/delete_contribution/<int:contribution_id>', methods=['POST'])
@login_required(roles=['president', 'secretary'])
def delete_contribution(contribution_id):
    """Allows President/Secretary to delete a pending contribution."""
    conn = get_db_connection()
    if conn is None:
        flash('Database connection error. Please try again later.', 'danger')
        return redirect(url_for('manage_contributions'))

    cursor = conn.cursor(buffered=True)

    try:
        # Optional: Check if the contribution is indeed pending before deleting
        cursor.execute("SELECT is_paid FROM contributions WHERE id = %s", (contribution_id,))
        contribution_status = cursor.fetchone()

        if not contribution_status:
            flash('Contribution not found.', 'danger')
            return redirect(url_for('manage_contributions'))

        if contribution_status[0]:  # is_paid is True
            flash('Approved contributions cannot be deleted.', 'danger')
            return redirect(url_for('manage_contributions'))

        cursor.execute("DELETE FROM contributions WHERE id = %s", (contribution_id,))
        conn.commit()
        flash('Contribution deleted successfully!', 'success')

    except mysql.connector.Error as err:
        flash(f'An error occurred while deleting contribution: {err}', 'danger')
        conn.rollback()
    finally:
        cursor.close()
        conn.close()
    return redirect(url_for('manage_contributions'))


# --- New Reports Feature ---
@app.route('/reports', methods=['GET', 'POST'])
@login_required(roles=['president', 'secretary'])
def reports():
    """Allows President/Secretary to generate and view various reports."""
    conn = get_db_connection()
    if conn is None:
        flash('Database connection error. Please try again later.', 'danger')
        return redirect(url_for('dashboard'))
    cursor = conn.cursor(dictionary=True, buffered=True)

    report_type = request.form.get('report_type')
    selected_month = request.form.get('month')
    selected_year = request.form.get('year')
    selected_member_id = request.form.get('member_id')

    report_data = []
    report_title = "Select a Report Type"
    report_headers = []

    # Fetch all members for the dropdown
    cursor.execute("SELECT id, name FROM users WHERE role IN ('member', 'president', 'secretary', 'treasurer') ORDER BY id")

    all_members = cursor.fetchall()

    # Fetch all years with data
    cursor.execute(
        "SELECT DISTINCT year FROM contributions UNION SELECT DISTINCT YEAR(start_date) FROM loans ORDER BY year DESC")
    all_years = [row['year'] for row in cursor.fetchall()]

    if request.method == 'POST':
        if report_type == 'monthly_contributions':
            if not selected_month or not selected_year:
                flash('Please select both month and year for Monthly Contributions report.', 'danger')
            else:
                report_title = f"Monthly Contributions Report - {datetime.strptime(selected_month, '%m').strftime('%B')} {selected_year}"
                report_headers = ["Member Name", "Amount", "Fine Amount", "Total Paid", "Status", "Payment Date",
                                  "UTR (Member)", "UTR (President)"]
                cursor.execute("""
                    SELECT u.name as member_name, c.amount, c.fine_amount, (c.amount + c.fine_amount) as total_paid,
                           c.is_paid, c.payment_date, c.utr_number, c.president_utr_number
                    FROM contributions c
                    JOIN users u ON c.user_id = u.id
                    WHERE c.month = %s AND c.year = %s
                    ORDER BY u.id
                """, (selected_month, selected_year))
                report_data = cursor.fetchall()

        elif report_type == 'yearly_contributions':
            if not selected_year:
                flash('Please select a year for Yearly Contributions report.', 'danger')
            else:
                report_title = f"Yearly Contributions Report - {selected_year}"
                report_headers = ["Member Name", "Total Contributions", "Total Fines", "Grand Total"]
                cursor.execute("""
                    SELECT u.name as member_name, SUM(c.amount) as total_amount, SUM(c.fine_amount) as total_fine_amount
                    FROM contributions c
                    JOIN users u ON c.user_id = u.id
                    WHERE c.year = %s AND c.is_paid = TRUE
                    GROUP BY u.id, u.name
                    ORDER BY u.id
                """, (selected_year,))
                report_data = cursor.fetchall()
                # Add a 'grand_total' field for display
                for row in report_data:
                    row['grand_total'] = (row['total_amount'] + row['total_fine_amount']).quantize(Decimal('0.01'),
                                                                                                   rounding=ROUND_HALF_UP)

        elif report_type == 'monthly_loan_interest':
            if not selected_month or not selected_year:
                flash('Please select both month and year for Monthly Loan Interest report.', 'danger')
            else:
                report_title = f"Monthly Loan Interest Collected - {datetime.strptime(selected_month, '%m').strftime('%B')} {selected_year}"
                report_headers = ["Borrower Name", "Loan Amount", "Interest Rate", "Interest Paid This Month",
                                  "Payment Date"]
                cursor.execute("""
                    SELECT u.name as borrower_name, l.amount as loan_amount, l.interest_rate, lp.interest_paid, lp.payment_date
                    FROM loan_payments lp
                    JOIN loans l ON lp.loan_id = l.id
                    JOIN users u ON l.user_id = u.id
                    WHERE MONTH(lp.payment_date) = %s AND YEAR(lp.payment_date) = %s
                    ORDER BY u.id, lp.payment_date
                """, (selected_month, selected_year))
                report_data = cursor.fetchall()

        elif report_type == 'yearly_loan_interest':
            if not selected_year:
                flash('Please select a year for Yearly Loan Interest report.', 'danger')
            else:
                report_title = f"Yearly Loan Interest Collected - {selected_year}"
                report_headers = ["Borrower Name", "Total Interest Paid in Year"]
                cursor.execute("""
                    SELECT u.name as borrower_name, SUM(lp.interest_paid) as total_interest_paid_yearly
                    FROM loan_payments lp
                    JOIN loans l ON lp.loan_id = l.id
                    JOIN users u ON l.user_id = u.id
                    WHERE YEAR(lp.payment_date) = %s
                    GROUP BY u.id, u.name
                    ORDER BY u.id
                """, (selected_year,))
                report_data = cursor.fetchall()

        elif report_type == 'member_contributions':
            if not selected_member_id:
                flash('Please select a member for Member Contributions report.', 'danger')
            else:
                cursor.execute("SELECT name FROM users WHERE id = %s", (selected_member_id,))
                member_name = cursor.fetchone()['name']
                report_title = f"Contributions History for {member_name}"
                report_headers = ["Month", "Year", "Amount", "Fine Amount", "Total Paid", "Status", "Payment Date"]
                cursor.execute("""
                    SELECT month, year, amount, fine_amount, (amount + fine_amount) as total_paid, is_paid, payment_date
                    FROM contributions
                    WHERE user_id = %s
                    ORDER BY year DESC, month DESC
                """, (selected_member_id,))
                report_data = cursor.fetchall()
                # Format month name for display
                for row in report_data:
                    row['month_name'] = datetime.strptime(str(row['month']), '%m').strftime('%B')

        elif report_type == 'member_loans':
            if not selected_member_id:
                flash('Please select a member for Member Loans report.', 'danger')
            else:
                cursor.execute("SELECT name FROM users WHERE id = %s", (selected_member_id,))
                member_name = cursor.fetchone()['name']
                report_title = f"Loan History for {member_name}"
                report_headers = ["Loan ID", "Amount", "Interest Rate", "Start Date", "Actual End Date", "Status",
                                  "Disbursement Type"]
                cursor.execute("""
                    SELECT id, amount, interest_rate, start_date, actual_end_date, status, disbursement_type, disbursement_details
                    FROM loans
                    WHERE user_id = %s
                    ORDER BY start_date DESC
                """, (selected_member_id,))
                report_data = cursor.fetchall()
                # Parse disbursement_details for display
                for row in report_data:
                    if row['disbursement_details']:
                        row['disbursement_details_parsed'] = json.loads(row['disbursement_details'])
                    else:
                        row['disbursement_details_parsed'] = {}

        elif report_type == 'all_members_summary':
            report_title = "All Members Summary"
            report_headers = ["Member Name", "Total Contributions", "Total Loans Taken", "Active Loans Count"]
            cursor.execute("""
                SELECT u.id, u.name,
                       SUM(CASE WHEN c.is_paid = TRUE THEN c.amount + c.fine_amount ELSE 0 END) as total_contributions,
                       SUM(CASE WHEN l.status IN ('approved', 'overdue', 'completed') THEN l.amount ELSE 0 END) as total_loans_taken,
                       COUNT(DISTINCT CASE WHEN l.status IN ('approved', 'overdue') THEN l.id ELSE NULL END) as active_loans_count
                FROM users u
                LEFT JOIN contributions c ON u.id = c.user_id
                LEFT JOIN loans l ON u.id = l.user_id
                WHERE u.role = 'member'
                GROUP BY u.id, u.name
                ORDER BY u.id
            """)
            report_data = cursor.fetchall()

    cursor.close()
    conn.close()

    return render_template('reports.html',
                           all_members=all_members,
                           all_years=all_years,
                           report_type=report_type,
                           selected_month=selected_month,
                           selected_year=selected_year,
                           selected_member_id=selected_member_id,
                           report_data=report_data,
                           report_title=report_title,
                           report_headers=report_headers,
                           excel_available=EXCEL_AVAILABLE,  # Pass availability to template
                           pdf_available=PDF_AVAILABLE)  # Pass availability to template


# --- Conceptual Export Routes (for demonstration, actual file generation may not work in sandbox) ---
@app.route('/export_report/<report_format>', methods=['POST'])
@login_required(roles=['president', 'secretary'])
def export_report(report_format):
    """
    Conceptual route for exporting reports.
    Actual file generation and download might not work directly in all sandbox environments.
    """
    report_type = request.form.get('report_type')
    selected_month = request.form.get('month')
    selected_year = request.form.get('year')
    selected_member_id = request.form.get('member_id')

    conn = get_db_connection()
    if conn is None:
        flash('Database connection error. Cannot generate report.', 'danger')
        return redirect(url_for('reports'))
    cursor = conn.cursor(dictionary=True, buffered=True)

    report_data = []
    report_title = "Report"
    report_headers = []

    # Re-fetch data based on report type, similar to the /reports GET/POST logic
    if report_type == 'monthly_contributions':
        report_title = f"Monthly Contributions - {datetime.strptime(selected_month, '%m').strftime('%B')} {selected_year}"
        report_headers = ["Member Name", "Amount", "Fine Amount", "Total Paid", "Status", "Payment Date",
                          "UTR (Member)", "UTR (President)"]
        cursor.execute("""
            SELECT u.name as member_name, c.amount, c.fine_amount, (c.amount + c.fine_amount) as total_paid,
                   c.is_paid, c.payment_date, c.utr_number, c.president_utr_number
            FROM contributions c
            JOIN users u ON c.user_id = u.id
            WHERE c.month = %s AND c.year = %s
            ORDER BY u.id
        """, (selected_month, selected_year))
        report_data = cursor.fetchall()
        # Convert Decimal and datetime objects to string for export
        for row in report_data:
            for key, value in row.items():
                if isinstance(value, Decimal):
                    row[key] = float(value)  # Convert Decimal to float for export
                elif isinstance(value, datetime):
                    row[key] = value.strftime('%Y-%m-%d %H:%M:%S')
                elif isinstance(value, date):
                    row[key] = value.strftime('%Y-%m-%d')
            row['is_paid'] = 'Paid' if row['is_paid'] else 'Pending'


    elif report_type == 'yearly_contributions':
        report_title = f"Yearly Contributions - {selected_year}"
        report_headers = ["Member Name", "Total Contributions", "Total Fines", "Grand Total"]
        cursor.execute("""
            SELECT u.name as member_name, SUM(c.amount) as total_amount, SUM(c.fine_amount) as total_fine_amount
            FROM contributions c
            JOIN users u ON c.user_id = u.id
            WHERE c.year = %s AND c.is_paid = TRUE
            GROUP BY u.id, u.name
            ORDER BY u.id
        """, (selected_year,))
        report_data = cursor.fetchall()
        for row in report_data:
            row['grand_total'] = (row['total_amount'] + row['total_fine_amount']).quantize(Decimal('0.01'),
                                                                                           rounding=ROUND_HALF_UP)
            for key, value in row.items():
                if isinstance(value, Decimal):
                    row[key] = float(value)

    elif report_type == 'monthly_loan_interest':
        report_title = f"Monthly Loan Interest Collected - {datetime.strptime(selected_month, '%m').strftime('%B')} {selected_year}"
        report_headers = ["Borrower Name", "Loan Amount", "Interest Rate", "Interest Paid This Month", "Payment Date"]
        cursor.execute("""
            SELECT u.name as borrower_name, l.amount as loan_amount, l.interest_rate, lp.interest_paid, lp.payment_date
            FROM loan_payments lp
            JOIN loans l ON lp.loan_id = l.id
            JOIN users u ON l.user_id = u.id
            WHERE MONTH(lp.payment_date) = %s AND YEAR(lp.payment_date) = %s
            ORDER BY u.id, lp.payment_date
        """, (selected_month, selected_year))
        report_data = cursor.fetchall()
        for row in report_data:
            for key, value in row.items():
                if isinstance(value, Decimal):
                    row[key] = float(value)
                elif isinstance(value, datetime):
                    row[key] = value.strftime('%Y-%m-%d %H:%M:%S')
                elif isinstance(value, date):
                    row[key] = value.strftime('%Y-%m-%d')

    elif report_type == 'yearly_loan_interest':
        report_title = f"Yearly Loan Interest Collected - {selected_year}"
        report_headers = ["Borrower Name", "Total Interest Paid in Year"]
        cursor.execute("""
            SELECT u.name as borrower_name, SUM(lp.interest_paid) as total_interest_paid_yearly
            FROM loan_payments lp
            JOIN loans l ON lp.loan_id = l.id
            JOIN users u ON l.user_id = u.id
            WHERE YEAR(lp.payment_date) = %s
            GROUP BY u.id, u.name
            ORDER BY u.id
        """, (selected_year,))
        report_data = cursor.fetchall()
        for row in report_data:
            for key, value in row.items():
                if isinstance(value, Decimal):
                    row[key] = float(value)

    elif report_type == 'member_contributions':
        cursor.execute("SELECT name FROM users WHERE id = %s", (selected_member_id,))
        member_name = cursor.fetchone()['name']
        report_title = f"Contributions History for {member_name}"
        report_headers = ["Month", "Year", "Amount", "Fine Amount", "Total Paid", "Status", "Payment Date"]
        cursor.execute("""
            SELECT month, year, amount, fine_amount, (amount + fine_amount) as total_paid, is_paid, payment_date
            FROM contributions
            WHERE user_id = %s
            ORDER BY year DESC, month DESC
        """, (selected_member_id,))
        report_data = cursor.fetchall()
        for row in report_data:
            row['month_name'] = datetime.strptime(str(row['month']), '%m').strftime('%B')
            for key, value in row.items():
                if isinstance(value, Decimal):
                    row[key] = float(value)
                elif isinstance(value, datetime):
                    row[key] = value.strftime('%Y-%m-%d %H:%M:%S')
                elif isinstance(value, date):
                    row[key] = value.strftime('%Y-%m-%d')
            row['is_paid'] = 'Paid' if row['is_paid'] else 'Pending'


    elif report_type == 'member_loans':
        cursor.execute("SELECT name FROM users WHERE id = %s", (selected_member_id,))
        member_name = cursor.fetchone()['name']
        report_title = f"Loan History for {member_name}"
        report_headers = ["Loan ID", "Amount", "Interest Rate", "Start Date", "Actual End Date", "Status",
                          "Disbursement Type", "Disbursement Details"]
        cursor.execute("""
            SELECT id, amount, interest_rate, start_date, actual_end_date, status, disbursement_type, disbursement_details
            FROM loans
            WHERE user_id = %s
            ORDER BY start_date DESC
        """, (selected_member_id,))
        report_data = cursor.fetchall()
        for row in report_data:
            if row['disbursement_details']:
                details = json.loads(row['disbursement_details'])
                if row['disbursement_type'] == 'cash':
                    row[
                        'disbursement_details_formatted'] = f"₹500: {details.get('notes_500', 0)}, ₹200: {details.get('notes_200', 0)}, ₹100: {details.get('notes_100', 0)}"
                elif row['disbursement_type'] == 'cheque':
                    row['disbursement_details_formatted'] = f"Cheque No: {details.get('cheque_number', 'N/A')}"
                elif row['disbursement_type'] == 'upi':
                    row['disbursement_details_formatted'] = f"UTR: {details.get('upi_utr', 'N/A')}"
                else:
                    row['disbursement_details_formatted'] = "N/A"
            else:
                row['disbursement_details_formatted'] = "N/A"

            for key, value in row.items():
                if isinstance(value, Decimal):
                    row[key] = float(value)
                elif isinstance(value, datetime):
                    row[key] = value.strftime('%Y-%m-%d %H:%M:%S')
                elif isinstance(value, date):
                    row[key] = value.strftime('%Y-%m-%d')

    elif report_type == 'all_members_summary':
        report_title = "All Members Summary"
        report_headers = ["Member Name", "Total Contributions", "Total Loans Taken", "Active Loans Count"]
        cursor.execute("""
            SELECT u.id, u.name,
                   SUM(CASE WHEN c.is_paid = TRUE THEN c.amount + c.fine_amount ELSE 0 END) as total_contributions,
                   SUM(CASE WHEN l.status IN ('approved', 'overdue', 'completed') THEN l.amount ELSE 0 END) as total_loans_taken,
                   COUNT(DISTINCT CASE WHEN l.status IN ('approved', 'overdue') THEN l.id ELSE NULL END) as active_loans_count
            FROM users u
            LEFT JOIN contributions c ON u.id = c.user_id
            LEFT JOIN loans l ON u.id = l.user_id
            WHERE u.role = 'member'
            GROUP BY u.id, u.name
            ORDER BY u.id
        """)
        report_data = cursor.fetchall()
        for row in report_data:
            for key, value in row.items():
                if isinstance(value, Decimal):
                    row[key] = float(value)

    cursor.close()
    conn.close()

    if not report_data:
        flash('No data found for the selected report criteria.', 'info')
        return redirect(url_for('reports'))

    if report_format == 'excel' and EXCEL_AVAILABLE:
        output = BytesIO()
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = report_title[:31]  # Max 31 chars for sheet title

        # Add title
        sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(report_headers))
        title_cell = sheet.cell(row=1, column=1, value=report_title)
        title_cell.font = Font(bold=True, size=16)
        title_cell.alignment = Alignment(horizontal='center', vertical='center')

        # Add headers
        for col_num, header in enumerate(report_headers, 1):
            cell = sheet.cell(row=3, column=col_num, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = Border(top=Side(style='thin'), bottom=Side(style='thin'), left=Side(style='thin'),
                                 right=Side(style='thin'))
            sheet.column_dimensions[get_column_letter(col_num)].width = 20  # Default width

        # Add data
        row_num = 4
        for row_dict in report_data:
            col_num = 1
            for header_key in report_headers:
                # Map header to dictionary key (e.g., "Member Name" -> "member_name")
                # This requires careful mapping or consistent naming.
                # For simplicity, we'll assume header names directly map to dictionary keys
                # or we'll need a more robust mapping logic.
                # Let's use a simplified mapping for common cases.
                key_map = {
                    "Member Name": "member_name",
                    "Borrower Name": "borrower_name",
                    "Loan ID": "id",
                    "Amount": "amount",
                    "Loan Amount": "loan_amount",
                    "Interest Rate (%)": "interest_rate",
                    "Start Date": "start_date",
                    "Actual End Date": "actual_end_date",
                    "Status": "status",
                    "Disbursement Type": "disbursement_type",
                    "Disbursement Details": "disbursement_details_formatted",  # Use formatted details
                    "Month": "month_name",  # For member_contributions
                    "Year": "year",
                    "Fine Amount": "fine_amount",
                    "Total Paid": "total_paid",
                    "UTR (Member)": "utr_number",
                    "UTR (President)": "president_utr_number",
                    "Total Contributions": "total_contributions",
                    "Total Fines": "total_fine_amount",
                    "Grand Total": "grand_total",
                    "Interest Paid This Month": "interest_paid",
                    "Total Interest Paid in Year": "total_interest_paid_yearly",
                    "Total Loans Taken": "total_loans_taken",
                    "Active Loans Count": "active_loans_count"
                }
                data_key = key_map.get(header_key, header_key.replace(" ", "_").lower())

                cell_value = row_dict.get(data_key, '')

                # Special handling for boolean status
                if data_key == 'is_paid':
                    cell_value = 'Paid' if cell_value else 'Pending'

                sheet.cell(row=row_num, column=col_num, value=cell_value)
                col_num += 1
            row_num += 1

        workbook.save(output)
        output.seek(0)
        return send_file(output, download_name=f"{report_title.replace(' ', '_')}.xlsx", as_attachment=True,
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    elif report_format == 'pdf' and PDF_AVAILABLE:
        output = BytesIO()
        doc = SimpleDocTemplate(output, pagesize=letter)
        styles = getSampleStyleSheet()

        elements = []

        # Title
        elements.append(Paragraph(report_title, styles['h1']))
        elements.append(Spacer(1, 12))

        # Table data
        table_data = [report_headers]
        for row_dict in report_data:
            row_values = []
            for header_key in report_headers:
                key_map = {
                    "Member Name": "member_name",
                    "Borrower Name": "borrower_name",
                    "Loan ID": "id",
                    "Amount": "amount",
                    "Loan Amount": "loan_amount",
                    "Interest Rate (%)": "interest_rate",
                    "Start Date": "start_date",
                    "Actual End Date": "actual_end_date",
                    "Status": "status",
                    "Disbursement Type": "disbursement_type",
                    "Disbursement Details": "disbursement_details_formatted",
                    "Month": "month_name",
                    "Year": "year",
                    "Fine Amount": "fine_amount",
                    "Total Paid": "total_paid",
                    "UTR (Member)": "utr_number",
                    "UTR (President)": "president_utr_number",
                    "Total Contributions": "total_contributions",
                    "Total Fines": "total_fine_amount",
                    "Grand Total": "grand_total",
                    "Interest Paid This Month": "interest_paid",
                    "Total Interest Paid in Year": "total_interest_paid_yearly",
                    "Total Loans Taken": "total_loans_taken",
                    "Active Loans Count": "active_loans_count"
                }
                data_key = key_map.get(header_key, header_key.replace(" ", "_").lower())

                cell_value = row_dict.get(data_key, '')

                if data_key == 'is_paid':
                    cell_value = 'Paid' if cell_value else 'Pending'
                elif isinstance(cell_value, Decimal):
                    cell_value = f"₹{float(cell_value):.2f}"
                elif isinstance(cell_value, datetime) or isinstance(cell_value, date):
                    cell_value = cell_value.strftime('%Y-%m-%d')

                row_values.append(str(cell_value))
            table_data.append(row_values)

        table = Table(table_data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#E0E0E0')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('BOX', (0, 0), (-1, -1), 0.5, colors.black),
        ]))
        elements.append(table)

        doc.build(output)
        output.seek(0)
        return send_file(output, download_name=f"{report_title.replace(' ', '_')}.pdf", as_attachment=True,
                         mimetype='application/pdf')

    else:
        flash('Report format not supported or required libraries not installed.', 'danger')
        return redirect(url_for('reports'))


# --- End New Reports Feature ---


if __name__ == '__main__':
    # Initial setup: Create a president user if none exists and ensure bank_balance entry exists
    conn = get_db_connection()
    if conn:
        cursor = conn.cursor(dictionary=True, buffered=True)

        # Check and create default president user
        cursor.execute("SELECT COUNT(*) FROM users WHERE role = 'president'")
        count = cursor.fetchone()['COUNT(*)']
        if count == 0:
            print("No president found. Creating a default president user.")
            try:
                hashed_password = generate_password_hash('password')
                cursor.execute(
                    "INSERT INTO users (name, username, email, contact_number, pan_number, aadhar_number, password, role) VALUES (%s, %s, %s, %s, %s, %s, %s, %s)",
                    ('President Admin', 'president', 'president@bachatgat.com', '9876543210', 'ABCDE1234F',
                     '123456789012', hashed_password, 'president')
                )
                conn.commit()
                print(
                    "Default president user created (username: president, password: password). Please change this password after first login.")
            except mysql.connector.Error as err:
                print(f"Error creating default president: {err}")
                conn.rollback()

        # Ensure bank_balance table has an entry and default_fine_amount, default_interest_rate columns
        try:
            cursor.execute("SELECT COUNT(*) FROM bank_balance")
            balance_count = cursor.fetchone()['COUNT(*)']
            if balance_count == 0:
                cursor.execute(
                    "INSERT INTO bank_balance (id, balance, default_fine_amount, default_interest_rate, default_contribution_amount, payment_start_day, payment_end_day) VALUES (1, %s, %s, %s, %s, %s, %s)",
                    (Decimal('0.00'), Decimal('50.00'), Decimal('10.00'), Decimal('100.00'), 1, 7))
                conn.commit()
                print(
                    "Initial bank_balance entry created with default fine amount, interest rate, contribution amount, and payment period.")
            else:
                # Basic migration checks for new columns
                cursor.execute("SHOW COLUMNS FROM bank_balance LIKE 'default_fine_amount'")
                if not cursor.fetchone():
                    cursor.execute(
                        "ALTER TABLE bank_balance ADD COLUMN default_fine_amount DECIMAL(10, 2) DEFAULT 50.00")
                    conn.commit()
                    print("Added default_fine_amount column to bank_balance table.")

                cursor.execute("SHOW COLUMNS FROM bank_balance LIKE 'default_interest_rate'")
                if not cursor.fetchone():
                    cursor.execute(
                        "ALTER TABLE bank_balance ADD COLUMN default_interest_rate DECIMAL(5, 2) DEFAULT 10.00")
                    conn.commit()
                    print("Added default_interest_rate column to bank_balance table.")

                cursor.execute("SHOW COLUMNS FROM bank_balance LIKE 'default_contribution_amount'")
                if not cursor.fetchone():
                    cursor.execute(
                        "ALTER TABLE bank_balance ADD COLUMN default_contribution_amount DECIMAL(10, 2) DEFAULT 100.00")
                    conn.commit()
                    print("Added default_contribution_amount column to bank_balance table.")

                cursor.execute("SHOW COLUMNS FROM bank_balance LIKE 'payment_start_day'")
                if not cursor.fetchone():
                    cursor.execute("ALTER TABLE bank_balance ADD COLUMN payment_start_day INT DEFAULT 1")
                    conn.commit()
                    print("Added payment_start_day column to bank_balance table.")

                cursor.execute("SHOW COLUMNS FROM bank_balance LIKE 'payment_end_day'")
                if not cursor.fetchone():
                    cursor.execute("ALTER TABLE bank_balance ADD COLUMN payment_end_day INT DEFAULT 7")
                    conn.commit()
                    print("Added payment_end_day column to bank_balance table.")

        except mysql.connector.Error as err:
            print(
                f"Error ensuring bank_balance table or default_fine_amount/interest_rate/contribution_amount/payment_days columns: {err}")
            conn.rollback()

        # Add actual_end_date column to loans table if it doesn't exist
        try:
            cursor.execute("SHOW COLUMNS FROM loans LIKE 'actual_end_date'")
            if not cursor.fetchone():
                cursor.execute("ALTER TABLE loans ADD COLUMN actual_end_date DATE NULL")
                conn.commit()
                print("Added actual_end_date column to loans table.")
        except mysql.connector.Error as err:
            print(f"Error adding actual_end_date column to loans table: {err}")
            conn.rollback()

        # New migration: Add UTR columns to contributions table
        try:
            cursor.execute("SHOW COLUMNS FROM contributions LIKE 'utr_number'")
            if not cursor.fetchone():
                cursor.execute("ALTER TABLE contributions ADD COLUMN utr_number VARCHAR(255) NULL")
                conn.commit()
                print("Added utr_number column to contributions table.")

            cursor.execute("SHOW COLUMNS FROM contributions LIKE 'president_utr_number'")
            if not cursor.fetchone():
                cursor.execute("ALTER TABLE contributions ADD COLUMN president_utr_number VARCHAR(255) NULL")
                conn.commit()
                print("Added president_utr_number column to contributions table.")

            # Ensure president_id exists for contributions table for who approved
            cursor.execute("SHOW COLUMNS FROM contributions LIKE 'president_id'")
            if not cursor.fetchone():
                cursor.execute("ALTER TABLE contributions ADD COLUMN president_id INT NULL")
                conn.commit()
                print("Added president_id column to contributions table.")
                # Add foreign key constraint if it doesn't exist (optional, but good practice)
                # This might fail if there's existing data that violates the constraint.
                try:
                    cursor.execute(
                        "ALTER TABLE contributions ADD CONSTRAINT fk_contributions_president FOREIGN KEY (president_id) REFERENCES users(id)")
                    conn.commit()
                    print("Added foreign key constraint for president_id in contributions table.")
                except mysql.connector.Error as fk_err:
                    print(
                        f"Warning: Could not add foreign key constraint for president_id in contributions table (might already exist or data mismatch): {fk_err}")


        except mysql.connector.Error as err:
            print(f"Error adding UTR columns or president_id to contributions table: {err}")
            conn.rollback()

        # Remove end_date column from loans table if it exists
        try:
            cursor.execute("SHOW COLUMNS FROM loans LIKE 'end_date'")
            if cursor.fetchone():
                cursor.execute("ALTER TABLE loans DROP COLUMN end_date")
                conn.commit()
                print("Removed end_date column from loans table.")
        except mysql.connector.Error as err:
            print(f"Error removing end_date column from loans table: {err}")
            conn.rollback()

        # --- New migration: Make president_id in loans table nullable ---
        try:
            cursor.execute("ALTER TABLE loans MODIFY COLUMN president_id INT NULL")
            conn.commit()
            print("Modified president_id column in loans table to be nullable.")
        except mysql.connector.Error as err:
            print(f"Error modifying president_id column in loans table to be nullable: {err}")
            conn.rollback()
        # --- End of new migration ---

        # --- New migration: Add disbursement_type and disbursement_details columns to loans table ---
        try:
            cursor.execute("SHOW COLUMNS FROM loans LIKE 'disbursement_type'")
            if not cursor.fetchone():
                cursor.execute("ALTER TABLE loans ADD COLUMN disbursement_type VARCHAR(50) NULL")
                conn.commit()
                print("Added disbursement_type column to loans table.")

            cursor.execute("SHOW COLUMNS FROM loans LIKE 'disbursement_details'")
            if not cursor.fetchone():
                cursor.execute("ALTER TABLE loans ADD COLUMN disbursement_details TEXT NULL")
                conn.commit()
                print("Added disbursement_details column to loans table.")
        except mysql.connector.Error as err:
            print(f"Error adding disbursement_type/details columns to loans table: {err}")
            conn.rollback()
        # --- End of new migration ---

        cursor.close()  # Close cursor after all operations
        conn.close()
    else:
        print("Could not connect to database for initial setup.")

    app.run(debug=True)  # Run in debug mode for development
